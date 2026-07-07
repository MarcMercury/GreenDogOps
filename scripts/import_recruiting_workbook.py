#!/usr/bin/env python3
"""Parse the updated Green Dog recruiting workbook (all tabs) and emit idempotent
SQL that MATCHES rows to existing people (by email, else unique name) and ADDS
candidates that are not yet in the system.

Improvements over the earlier one-shot import:
  * Reads the real Google-Doc / Drive HYPERLINKS out of each cell (the earlier
    import only captured the visible cell text, so document links were lost and
    `resume_url` ended up holding the "Found on" source instead of a URL).
  * Enriches existing recruiting rows instead of wiping + reinserting, so manual
    edits and non-workbook candidates are preserved.
  * Collects every attached document (interview notes, resume PDF, CSR exercise)
    into a labelled "Documents" block and puts the primary doc link in
    `resume_url`.

Usage:
    python3 scripts/import_recruiting_workbook.py \
        --xlsx /tmp/recruiting.xlsx \
        --db-json /tmp/db_persons.json \
        --out scripts/recruiting_workbook_import.sql
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import date, datetime

import openpyxl

PLACEHOLDER_EMAILS = {
    "need", "pending", "pending...", "pending......", "pending.....",
    "n/a", "na", "tbd", "none", "nan", "",
}


# --------------------------------------------------------------------------- #
# SQL literal helpers
# --------------------------------------------------------------------------- #
def s(v):
    """SQL string literal (or NULL)."""
    if v is None:
        return "null"
    v = str(v).strip()
    if v == "" or v.lower() in ("none", "nan"):
        return "null"
    return "'" + v.replace("'", "''") + "'"


def num(v):
    if v is None or str(v).strip() == "":
        return "null"
    try:
        f = float(v)
    except (ValueError, TypeError):
        return "null"
    # Score of 0 in the sheet means "unrated" -> leave null.
    if f == 0:
        return "null"
    return str(f)


def boolean(v):
    if v is None:
        return "null"
    t = str(v).strip().lower()
    if t in ("yes", "true", "1", "y"):
        return "true"
    if t in ("no", "false", "0", "n"):
        return "false"
    return "null"


def sql_date(v):
    if v is None:
        return "null"
    if isinstance(v, datetime):
        return "'" + v.strftime("%Y-%m-%d") + "'"
    if isinstance(v, date):
        return "'" + v.strftime("%Y-%m-%d") + "'"
    raw = str(v).strip()
    if not raw:
        return "null"
    # Pull the first date-looking token out of messy cells like "5/15/2026 @ 1:30".
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", raw)
    if m:
        mo, da, yr = m.groups()
        yr = int(yr)
        if yr < 100:
            yr += 2000
        try:
            return "'" + date(yr, int(mo), int(da)).strftime("%Y-%m-%d") + "'"
        except ValueError:
            return "null"
    for fmt in ("%Y-%m-%d", "%b %d, %Y", "%B %d, %Y"):
        try:
            return "'" + datetime.strptime(raw, fmt).strftime("%Y-%m-%d") + "'"
        except ValueError:
            continue
    return "null"


# --------------------------------------------------------------------------- #
# Text helpers
# --------------------------------------------------------------------------- #
def clean_text(v):
    if v is None:
        return None
    t = re.sub(r"\s+", " ", str(v)).strip()
    return t or None


def strip_title(v):
    if not v:
        return v
    return re.sub(r"^(dr|mr|mrs|ms)\.?\s+", "", str(v).strip(), flags=re.IGNORECASE)


def clean_email(v):
    if v is None:
        return None
    t = str(v).strip().strip(".,;:| ").lower()
    if t in PLACEHOLDER_EMAILS:
        return None
    if " " in t or "@" not in t:
        return None
    # Handle "a@b.com <mailto:...>" leftovers.
    t = t.split()[0].strip("<>")
    if "@" not in t or t.startswith("http"):
        return None
    return t


def clean_phone(v):
    if v is None:
        return None
    t = re.sub(r"\s+", " ", str(v)).strip()
    # Skip obvious non-phones ("Team 3", "Blueragg700@gmail.com").
    if "@" in t or not re.search(r"\d", t):
        return None
    # Numeric cells arrive as floats -> drop the trailing ".0".
    t = re.sub(r"\.0$", "", t)
    return t or None


def norm_name(v):
    """Normalise a name for matching: drop titles, nicknames, punctuation."""
    if not v:
        return ""
    t = str(v).lower()
    t = re.sub(r"\(.*?\)", " ", t)          # (Niko Alzate)
    t = re.sub(r"[\u2018\u2019\"']", " ", t)  # smart quotes / nickname quotes
    t = re.sub(r"\bdr\.?\b", " ", t)         # Dr.
    t = re.sub(r"[^a-z ]", " ", t)           # punctuation / digits
    t = re.sub(r"\s+", " ", t).strip()
    return t


def split_name(full):
    full = clean_text(full) or ""
    full = full.strip("'\" ")
    full = re.sub(r"\s+\|+\s*$", "", full)   # trailing "|" artefacts
    full = strip_title(full)
    if not full:
        return None, None
    parts = full.split(" ")
    if len(parts) == 1:
        return parts[0], None
    return parts[0], " ".join(parts[1:])


def cell_link(cell):
    """Return (text, url) for a cell, resolving hyperlinks and inline URLs."""
    val = clean_text(cell.value)
    url = None
    if cell.hyperlink and cell.hyperlink.target:
        url = cell.hyperlink.target
    elif val and val.lower().startswith("http"):
        url = val.split()[0]
    if url and url.lower().startswith("mailto:"):
        url = None
    return val, url


def looks_like_person(name):
    if not name:
        return False
    low = name.strip().lower()
    if low in ("team 3", "team 4", "team 5", "others", "my pet"):
        return False
    if "http" in low or "@" in low:
        return False
    # All-caps short label with no lowercase == section divider ("VET TECH").
    if name.isupper() and len(name.split()) <= 3 and not any(c.islower() for c in name):
        return False
    return True


# --------------------------------------------------------------------------- #
# Workbook parsing
# --------------------------------------------------------------------------- #
class Candidate:
    __slots__ = (
        "first_name", "last_name", "full_name", "email", "phone",
        "target_title", "pipeline", "stage", "source", "interview_date",
        "score", "follow_up_date", "keep_for_future", "status_notes",
        "resume_url", "docs",
    )

    def __init__(self):
        self.first_name = self.last_name = self.full_name = None
        self.email = self.phone = None
        self.target_title = self.pipeline = self.stage = self.source = None
        self.interview_date = self.follow_up_date = None
        self.score = None
        self.keep_for_future = None
        self.status_notes = None
        self.resume_url = None
        self.docs = []   # list of (label, url)

    def add_doc(self, label, url):
        if not url:
            return
        for _, u in self.docs:
            if u == url:
                return
        self.docs.append((clean_text(label) or "Document", url))


def pick_resume(cand: Candidate):
    """Choose the primary document link for resume_url."""
    if not cand.docs:
        return None
    # Prefer an actual resume PDF, else the interview/notes doc, else first.
    for label, url in cand.docs:
        lab = (label or "").lower()
        if "resume" in lab or url.lower().endswith(".pdf") or "drive.google" in url:
            return url
    return cand.docs[0][1]


def docs_block(cand: Candidate):
    if not cand.docs:
        return None
    lines = ["Documents:"]
    for label, url in cand.docs:
        lines.append(f"- {label}: {url}")
    return "\n".join(lines)


# Column layouts keyed by the fields we need. Index = column position.
LAYOUTS = {
    "All In House Positions": dict(
        header=1, name=0, target=1, status_notes=2, stage=3, email=4, phone=5,
        interview=6, score=7, doc_cols=[8, 9], source=10, keep=11, follow=12,
        pipeline="All In House Positions",
    ),
    "Remote CSR": dict(
        header=1, name=0, source=1, status_notes=2, stage=3, interview=4,
        email=5, phone=6, score=7, doc_cols=[8, 9, 10], keep=11, follow=12,
        pipeline="Remote CSR", target="Remote CSR",
    ),
    "DVM Vet America ": dict(
        header=0, name=0, status_notes=1, stage=2, interview=3, email=4,
        phone=5, score=6, doc_cols=[7], source=8, keep=9, follow=10,
        pipeline="DVM Vet America", target="DVM",
    ),
    "Volunteers": dict(
        header=0, name=0, status_notes=1, stage=2, email=3, phone=4, score=5,
        doc_cols=[6], source=7, keep=8,
        pipeline="Volunteers, Externs", target="Volunteer",
    ),
    "Externs": dict(
        header=0, name=0, status_notes=1, stage=2, email=3, phone=4, score=5,
        doc_cols=[6], source=7, keep=8,
        pipeline="Volunteers, Externs", target="Extern",
    ),
}


def parse_standard(ws, cfg):
    out = []
    rows = list(ws.iter_rows())
    start = cfg["header"] + 1
    cur_target = cfg.get("target")
    for row in rows[start:]:
        def cv(key):
            ci = cfg.get(key)
            if ci is None or ci >= len(row):
                return None
            return clean_text(row[ci].value)

        name = cv("name")
        # Divider row that names a position group (e.g. "VET TECH || VET TECH").
        target_cell = None
        if "target" in cfg and isinstance(cfg["target"], int):
            target_cell = cv("target")
        if name and target_cell and name.strip() == target_cell.strip() \
                and not any(c.islower() for c in name):
            cur_target = target_cell
            continue
        if not looks_like_person(name):
            # Track running position group header if present.
            if name and name.isupper():
                cur_target = name
            continue

        c = Candidate()
        fn, ln = split_name(name)
        c.first_name, c.last_name = fn, ln
        c.full_name = " ".join(x for x in [fn, ln] if x)
        c.email = clean_email(row[cfg["email"]].value) if cfg.get("email") is not None and cfg["email"] < len(row) else None
        c.phone = clean_phone(row[cfg["phone"]].value) if cfg.get("phone") is not None and cfg["phone"] < len(row) else None
        c.stage = cv("stage")
        c.status_notes = cv("status_notes")
        c.score = row[cfg["score"]].value if cfg.get("score") is not None and cfg["score"] < len(row) else None
        c.keep_for_future = cv("keep")
        c.pipeline = cfg.get("pipeline")

        if isinstance(cfg.get("target"), int):
            c.target_title = target_cell or cur_target
        else:
            c.target_title = cfg.get("target") or cur_target

        if cfg.get("source") is not None and cfg["source"] < len(row):
            c.source = clean_text(row[cfg["source"]].value)

        if cfg.get("interview") is not None and cfg["interview"] < len(row):
            c.interview_date = row[cfg["interview"]].value
        if cfg.get("follow") is not None and cfg["follow"] < len(row):
            c.follow_up_date = row[cfg["follow"]].value

        for ci in cfg.get("doc_cols", []):
            if ci < len(row):
                label, url = cell_link(row[ci])
                c.add_doc(label, url)
        out.append(c)
    return out


def parse_hired(ws):
    out = []
    rows = list(ws.iter_rows())
    for row in rows[1:]:
        def cv(ci):
            return clean_text(row[ci].value) if ci < len(row) else None

        notes = cv(0)
        fn = cv(2)
        ln = cv(3)
        if not fn and not ln:
            continue
        combined = f"{fn or ''} {ln or ''}".strip()
        if not looks_like_person(combined) or combined.lower() == "all hired & offers made":
            continue
        if not ln and fn:
            fn, ln = split_name(fn)
        fn = strip_title(fn)
        c = Candidate()
        c.first_name, c.last_name = fn, ln
        c.full_name = " ".join(x for x in [fn, ln] if x)
        c.email = clean_email(row[4].value) if len(row) > 4 else None
        c.phone = clean_phone(row[5].value) if len(row) > 5 else None
        c.target_title = cv(6)
        c.source = cv(1)
        c.status_notes = notes
        c.stage = "Hired"
        c.pipeline = "Hired"
        loc = cv(8)
        if loc:
            c.status_notes = (c.status_notes + " | " if c.status_notes else "") + f"location: {loc}"
        for ci in (9, 10):
            if ci < len(row):
                label, url = cell_link(row[ci])
                c.add_doc(label, url)
        out.append(c)
    return out


def parse_old_dvm(ws):
    """Legacy international DVM sheet: first/last split, many divider rows."""
    out = []
    rows = list(ws.iter_rows())
    for row in rows[2:]:
        def cv(ci):
            return clean_text(row[ci].value) if ci < len(row) else None

        fn = cv(2)
        ln = cv(3)
        notes = cv(0)
        if not fn:
            continue
        if "http" in fn.lower() or "@" in fn:
            continue
        combined = f"{fn} {ln or ''}".strip()
        if not looks_like_person(combined):
            continue
        c = Candidate()
        # Some cells carry the whole name in the first-name slot.
        if not ln:
            fn, ln = split_name(fn)
        fn = strip_title(fn)
        c.first_name, c.last_name = fn, ln
        c.full_name = " ".join(x for x in [fn, ln] if x)
        c.email = clean_email(row[4].value) if len(row) > 4 else None
        c.phone = clean_phone(row[5].value) if len(row) > 5 else None
        c.target_title = cv(6) or "DVM"
        c.source = cv(1)
        c.status_notes = notes
        c.pipeline = "DVM Vet America"
        loc = cv(8)
        if loc:
            c.status_notes = (c.status_notes + " | " if c.status_notes else "") + f"location: {loc}"
        for ci in range(len(row)):
            label, url = cell_link(row[ci])
            if url:
                c.add_doc(label, url)
        out.append(c)
    return out


# --------------------------------------------------------------------------- #
# Dedupe + merge across tabs
# --------------------------------------------------------------------------- #
def merge_into(dst: Candidate, src: Candidate):
    def pref(a, b):
        return a if a else b
    dst.email = pref(dst.email, src.email)
    dst.phone = pref(dst.phone, src.phone)
    dst.stage = pref(dst.stage, src.stage)
    dst.source = pref(dst.source, src.source)
    dst.score = dst.score if (dst.score and float_or0(dst.score) > 0) else src.score
    dst.interview_date = pref(dst.interview_date, src.interview_date)
    dst.follow_up_date = pref(dst.follow_up_date, src.follow_up_date)
    dst.keep_for_future = pref(dst.keep_for_future, src.keep_for_future)
    dst.status_notes = pref(dst.status_notes, src.status_notes)
    dst.target_title = pref(dst.target_title, src.target_title)
    if src.pipeline and src.pipeline not in (dst.pipeline or ""):
        dst.pipeline = (dst.pipeline + ", " if dst.pipeline else "") + src.pipeline
    for label, url in src.docs:
        dst.add_doc(label, url)


def float_or0(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def dedupe(cands):
    records = []
    email_idx = {}
    name_idx = {}
    for c in cands:
        nm = norm_name(c.full_name)
        if not c.email and not nm:
            continue
        target = None
        if c.email and c.email in email_idx:
            target = email_idx[c.email]
        elif nm and nm in name_idx:
            other = name_idx[nm]
            # Merge by name only when it can't conflate two distinct emails.
            if not c.email or not other.email or other.email == c.email:
                target = other
        if target:
            merge_into(target, c)
        else:
            records.append(c)
            target = c
        if target.email:
            email_idx[target.email] = target
        if nm:
            name_idx[nm] = target
    return records


# --------------------------------------------------------------------------- #
# DB matching
# --------------------------------------------------------------------------- #
def load_db(path):
    people = json.load(open(path))
    by_email = {}
    by_name = {}
    for p in people:
        em = (p.get("email") or "").strip().lower()
        if em:
            by_email.setdefault(em, p)
        full = p.get("full_name") or " ".join(
            x for x in [p.get("first_name"), p.get("last_name")] if x
        )
        nm = norm_name(full)
        if nm:
            by_name.setdefault(nm, []).append(p)
    return by_email, by_name


def match(cand, by_email, by_name):
    if cand.email and cand.email in by_email:
        return by_email[cand.email], "email"
    nm = norm_name(cand.full_name)
    if nm and nm in by_name:
        people = by_name[nm]
        applicants = [p for p in people if p["status"] == "applicant"]
        pool = applicants or people
        if len(pool) == 1:
            return pool[0], "name"
        # Ambiguous: enrich the first applicant rather than create a duplicate.
        return pool[0], "name-ambiguous"
    return None, None


# --------------------------------------------------------------------------- #
# SQL emission
# --------------------------------------------------------------------------- #
def emit(cands, by_email, by_name):
    lines = [
        "set search_path = greendogops, public;",
        "begin;",
        "-- Recruiting workbook import: match + enrich existing people, add new.",
    ]
    matched = new = 0
    for c in cands:
        resume = pick_resume(c)
        notes = docs_block(c)
        db, how = match(c, by_email, by_name)
        if db:
            matched += 1
            pid = db["id"]
            existing_resume = db.get("resume_url")
            # Prefer the workbook doc URL; keep an existing *real* URL otherwise;
            # discard the earlier garbled non-URL values.
            final_resume = resume
            if not final_resume and existing_resume and str(existing_resume).lower().startswith("http"):
                final_resume = existing_resume
            lines.append(
                "update greendogops.person set "
                f"email = coalesce(email, {s(c.email)}), "
                f"phone_mobile = coalesce(phone_mobile, {s(c.phone)}) "
                f"where id = '{pid}';"
            )
            lines.append(
                "insert into greendogops.person_recruiting as r "
                "(person_id, target_title, pipeline, stage, source, interview_date, "
                " score, resume_url, keep_for_future, follow_up_date, status_notes, notes) "
                f"values ('{pid}', {s(c.target_title)}, {s(c.pipeline)}, {s(c.stage)}, "
                f"{s(c.source)}, {sql_date(c.interview_date)}, {num(c.score)}, "
                f"{s(final_resume)}, {boolean(c.keep_for_future)}, {sql_date(c.follow_up_date)}, "
                f"{s(c.status_notes)}, {s(notes)}) "
                "on conflict (person_id) do update set "
                "target_title   = coalesce(excluded.target_title, r.target_title), "
                "pipeline       = coalesce(excluded.pipeline, r.pipeline), "
                "stage          = coalesce(excluded.stage, r.stage), "
                "source         = coalesce(excluded.source, r.source), "
                "interview_date = coalesce(excluded.interview_date, r.interview_date), "
                "score          = coalesce(excluded.score, r.score), "
                # resume_url is set outright: excluded already carries the best
                # value (workbook link > existing real URL) and this clears the
                # earlier garbled non-URL values.
                "resume_url     = excluded.resume_url, "
                "keep_for_future= coalesce(excluded.keep_for_future, r.keep_for_future), "
                "follow_up_date = coalesce(excluded.follow_up_date, r.follow_up_date), "
                "status_notes   = coalesce(excluded.status_notes, r.status_notes), "
                "notes          = coalesce(excluded.notes, r.notes);"
            )
        else:
            new += 1
            full = c.full_name or " ".join(x for x in [c.first_name, c.last_name] if x)
            lines.append(
                "with p as (insert into greendogops.person "
                "(status, first_name, last_name, full_name, email, phone_mobile) values ("
                f"'applicant', {s(c.first_name)}, {s(c.last_name)}, {s(full)}, "
                f"{s(c.email)}, {s(c.phone)}) returning id) "
                "insert into greendogops.person_recruiting "
                "(person_id, target_title, pipeline, stage, source, interview_date, "
                " score, resume_url, keep_for_future, follow_up_date, status_notes, notes) "
                f"select id, {s(c.target_title)}, {s(c.pipeline)}, {s(c.stage)}, {s(c.source)}, "
                f"{sql_date(c.interview_date)}, {num(c.score)}, {s(resume)}, "
                f"{boolean(c.keep_for_future)}, {sql_date(c.follow_up_date)}, "
                f"{s(c.status_notes)}, {s(notes)} from p;"
            )
    lines.append("commit;")
    lines.append(
        "select count(*) as applicants from greendogops.person where status='applicant';"
    )
    sys.stderr.write(f"candidates={len(cands)} matched={matched} new={new}\n")
    return "\n".join(lines) + "\n"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="/tmp/recruiting.xlsx")
    ap.add_argument("--db-json", default="/tmp/db_persons.json")
    ap.add_argument("--out", default="scripts/recruiting_workbook_import.sql")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    cands = []
    for title, cfg in LAYOUTS.items():
        if title in wb.sheetnames:
            cands += parse_standard(wb[title], cfg)
    if "HIRED " in wb.sheetnames:
        cands += parse_hired(wb["HIRED "])
    if "OLD!!! ALL DVM CANDIDATES" in wb.sheetnames:
        cands += parse_old_dvm(wb["OLD!!! ALL DVM CANDIDATES"])

    cands = dedupe(cands)
    by_email, by_name = load_db(args.db_json)
    sql = emit(cands, by_email, by_name)
    with open(args.out, "w") as f:
        f.write(sql)
    sys.stderr.write(f"wrote {args.out}\n")


if __name__ == "__main__":
    main()
