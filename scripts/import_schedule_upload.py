#!/usr/bin/env python3
"""Import the "Schedule Upload GddOPs.xlsx" workbook into the scheduling / HR
tables. Three data feeds, one reviewable SQL file:

  1. CALLOUT 26   -> attendance markings on sched_assignment.
        Key: T=late, TE=late_excused, A=absent, AE=absent_excused.
        Blue font = "did not report to Angie", black = "reported to Angie"
        (captured in attendance_note). When an employee has NO scheduled shift
        that day, a minimal assignment is CREATED (on their dominant role/line
        and location) so the marking always surfaces on the HR Attendance tab.

  2. VACATION CALENDAR 26 -> honoured PTO. Each off-day becomes a
        person_pto_day row AND consecutive runs collapse into an approved
        person_time_off range (kind=pto, status=approved). Existing scheduled
        assignments on those days are flipped to attendance_status='pto'.
        (ALT CAL tab is intentionally ignored.)

  3. Month tabs (January..August) -> schedule placements. Every grid cell that
        resolves to a roster person is OVERLAID onto that week (weeks/lines are
        created if missing; existing assignments are never wiped or duplicated).

People are matched conservatively against greendogops.person (employee /
contractor / former). Anything ambiguous or unmatched is skipped and reported.

OUTPUT (review, then apply):
  * .data/schedule_upload.sql         - idempotent, transactional-per-block SQL
  * .data/schedule_upload_report.txt  - coverage + unmatched names

Apply with:  ./scripts/supabase-sql.sh -f .data/schedule_upload.sql
"""
import re
import sys
from collections import defaultdict, Counter
from datetime import date, timedelta
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from match_schedule_staff import (  # noqa: E402
    DR_PREFIX, norm, sql_str, run_sql, load_people, match_person,
)
from import_schedule_weeks import classify, resolve_status  # noqa: E402

XLSX = "public/Schedule Upload GddOPs.xlsx"
OUT_SQL = ".data/schedule_upload.sql"
OUT_REPORT = ".data/schedule_upload_report.txt"
YEAR = 2026
# Schedule is intentionally blank from this date on (scheduler's clean slate);
# never (re)create placements or callout shifts on/after it.
CLEAR_FROM = date(2026, 9, 1)

MONTH_TABS = {
    "January": 1, "FEB": 2, "MAR": 3, "April": 4,
    "May": 5, "June": 6, "July": 7, "August": 8,
}

# Location header text (in the month grids) -> DB location short_code.
LOC_HEADER = {
    "SO": "SO", "SHERMAN": "SO",
    "VENICE": "VEN", "VEN": "VEN",
    "AETNA": "VAN", "VAN NUYS": "VAN", "VAN": "VAN", "VALLEY": "VAN",
    "MPMV": "MPMV",
}
DAY_NAMES = {
    "sunday": 0, "monday": 1, "tuesday": 2, "wednesday": 3,
    "thursday": 4, "friday": 5, "saturday": 6,
}
MONTH_NAMES = {
    "january": 1, "february": 2, "march": 3, "april": 4, "may": 5, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10, "november": 11,
    "december": 12,
}
# CALLOUT marking -> attendance_status enum.
ATT_MAP = {"T": "late", "TE": "late_excused", "A": "absent", "AE": "absent_excused"}

# Grid role labels that classify() emits but that don't exist verbatim as a
# shift-template role -> the real template role name.
ROLE_FIX = {
    "surgery tech 1": "Surgery Tech",
    "in house admin marketing assist": "Admin/Mrkt Asst.",
}

# Department-header rows ("VET-SURGERY", "VET-AP", ...) carry each department's
# DVM names. classify() in import_schedule_weeks keys on hyphenated tokens, but
# norm() converts the hyphen to a space, so those headers never match there.
# Detect them here (hyphen-free) so the running department + its doctors resolve.
DEPT_HEAD = [
    ("vet surgery", "SURG"), ("vet ap", "AP"), ("vet nad", "NAD"),
    ("vet im", "IM"), ("vet exotic", "EXO"), ("vet mpmv", "MPMV"),
    ("vet cardio", "CARD"),
]


def classify2(label, current):
    """classify() with robust VET-<dept> header detection."""
    low = norm(label)
    for kw, dept in DEPT_HEAD:
        if kw in low:
            return dept, "DVM", dept
    return classify(label, current)


def clean_name(frag):
    """Strip a trailing shift-marker digit ('Laura Lucia1' -> 'Laura Lucia')
    and surrounding punctuation the grid appends to duplicate placements."""
    frag = re.sub(r"\s*\d+\s*$", "", frag.strip())
    return frag.strip(" .-*")


def _lev1(a, b):
    """True if strings are within edit distance 1 (cheap, for typo tolerance)."""
    if a == b:
        return True
    la, lb = len(a), len(b)
    if abs(la - lb) > 1:
        return False
    if la == lb:
        return sum(x != y for x, y in zip(a, b)) == 1
    if la > lb:
        a, b, la, lb = b, a, lb, la
    for i in range(la + 1):
        if a[:i] + a[i:] and a[:i] == b[:i] and a[i:] == b[i + 1:]:
            return True
    return False


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return ""
    return str(v).strip()


def daynum(ws, r, c):
    """Parse a calendar day-number cell ('4', '4.0', '31') -> int or None."""
    v = ws.cell(row=r, column=c).value
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def wrap_month(m):
    """Return (year_offset, month 1..12) for an out-of-range month int."""
    if m < 1:
        return -1, m + 12
    if m > 12:
        return 1, m - 12
    return 0, m


def reconstruct_week(nums, month, year=YEAR):
    """Given 7 Sun..Sat day-numbers (some None) and the tab's nominal month,
    return the 7 real dates. Tries the nominal month and its neighbours so a
    week that spills across a month boundary still resolves."""
    i0 = next((i for i, n in enumerate(nums) if n is not None), None)
    if i0 is None:
        return None
    for cand in (month, month - 1, month + 1):
        yo, mm = wrap_month(cand)
        try:
            anchor = date(year + yo, mm, nums[i0])
        except ValueError:
            continue
        sunday = anchor - timedelta(days=i0)
        ok = all(
            nums[j] is None or (sunday + timedelta(days=j)).day == nums[j]
            for j in range(7)
        )
        if ok:
            return [sunday + timedelta(days=j) for j in range(7)]
    return None


def sunday_of(d):
    """Sunday (0=Sun sheet convention) that starts d's week."""
    return d - timedelta(days=(d.weekday() + 1) % 7)


def sheet_dow(d):
    return (d.weekday() + 1) % 7


# ---------------------------------------------------------------------------
# People matching (adds nickname fallbacks over match_schedule_staff)
# ---------------------------------------------------------------------------

class Matcher:
    def __init__(self):
        self.active = load_people(("employee", "contractor"))
        self.allp = load_people(("employee", "contractor", "former"))
        self.idx = self._index(self.active)
        self.cache = {}
        self.unmatched = Counter()

    @staticmethod
    def _index(people):
        idx = {"first": defaultdict(set), "pref": defaultdict(set),
               "last": defaultdict(set)}
        for p in people:
            if p["first"]:
                idx["first"][p["first"]].add(p["id"])
            if p["pref"]:
                idx["pref"][p["pref"]].add(p["id"])
            if p["last"]:
                idx["last"][p["last"]].add(p["id"])
        return idx

    def _by_id(self, pid):
        for p in self.active:
            if p["id"] == pid:
                return p
        return None

    def resolve(self, raw):
        raw = clean_name(raw or "")
        key = norm(DR_PREFIX.sub("", raw))
        if len(key) < 2:
            return None
        if key in self.cache:
            return self.cache[key]

        person, _ = match_person(raw, self.active)
        if not person:
            person, _ = match_person(raw, self.allp)
        pid = person["id"] if person else None

        if not pid:
            toks = [t for t in key.split() if len(t) > 1]
            # Single-token nickname -> unique first / pref / last hit.
            if len(toks) == 1:
                t = toks[0]
                for field in ("first", "pref", "last"):
                    ids = self.idx[field].get(t)
                    if ids and len(ids) == 1:
                        pid = next(iter(ids))
                        break
                # Prefix match (e.g. "vero" -> "veronica").
                if not pid and len(t) >= 3:
                    hits = {
                        i for field in ("first", "pref")
                        for name, ids in self.idx[field].items()
                        if name.startswith(t) for i in ids
                    }
                    if len(hits) == 1:
                        pid = next(iter(hits))
            # First name exact + last name within one typo (roster misspellings).
            elif len(toks) >= 2:
                first, last = toks[0], toks[-1]
                hits = {
                    p["id"] for p in self.active
                    if (p["first"] == first or p["pref"] == first)
                    and p["last"] and _lev1(p["last"], last)
                }
                if len(hits) == 1:
                    pid = next(iter(hits))

        if not pid:
            self.unmatched[raw.strip()] += 1
        self.cache[key] = pid
        return pid


# ---------------------------------------------------------------------------
# Line resolution (templates -> per-week snapshot line via template_id)
# ---------------------------------------------------------------------------

def load_templates():
    rows = run_sql(
        "set search_path=greendogops,public; "
        "select t.id, coalesce(d.code,'') dept, coalesce(r.name,t.label) role, "
        "t.sort_order from sched_shift_template t "
        "join sched_department d on d.id=t.department_id "
        "left join sched_role r on r.id=t.role_id "
        "where t.is_active order by t.sort_order;"
    )
    by_dr = defaultdict(list)
    for t in rows:
        by_dr[(t["dept"], norm(t["role"]))].append(t["id"])
    return by_dr


def load_locations():
    return {r["short_code"]: r["id"] for r in run_sql(
        "set search_path=greendogops,public; "
        "select id, short_code from location where is_active;")}


def load_dvm_ids():
    return {r["id"] for r in run_sql(
        "set search_path=greendogops,public; select id from person "
        "where status in ('employee','contractor','former') "
        "and full_name ilike 'dr.%';")}


# ---------------------------------------------------------------------------
# Month-tab schedule parsing
# ---------------------------------------------------------------------------

def build_day_columns(ws, header_row):
    """From a WEEK header row return ordered [(col, day_idx, loc_short)] plus the
    date-number row and location sub-header row indices."""
    date_row = header_row + 1
    loc_row = header_row + 2
    # Day-name start columns.
    day_cols = []
    for c in range(1, ws.max_column + 1):
        v = norm(cell(ws, header_row, c))
        if v in DAY_NAMES:
            day_cols.append((c, DAY_NAMES[v]))
    if not day_cols:
        return None
    day_cols.sort()
    starts = [c for c, _ in day_cols]

    def day_for(col):
        d = None
        for (c, idx) in day_cols:
            if col >= c:
                d = idx
            else:
                break
        return d

    # Location sub-headers within the day spans (skip summary cols to the right).
    last_day_col = starts[-1]
    right_bound = last_day_col + 8
    cols = []
    seen_first_day = {}
    for c in range(starts[0], min(right_bound, ws.max_column) + 1):
        lv = cell(ws, loc_row, c).upper()
        sc = LOC_HEADER.get(lv)
        if sc:
            cols.append((c, day_for(c), sc))
    # Sunday (day 0) usually has only a date cell, no location header -> its
    # start column is the single SO clinic.
    for (c, idx) in day_cols:
        if idx == 0 and not any(cc == c for cc, _, _ in cols):
            cols.append((c, 0, "SO"))
    cols.sort()
    return {"cols": cols, "date_row": date_row, "day_cols": day_cols}


def parse_month_tab(ws, month, matcher, by_dr, dvm_ids, loc_ids,
                    placements, dominant_role, dominant_loc, exceptions):
    r = 1
    max_r = ws.max_column and ws.max_row
    while r <= max_r:
        if norm(cell(ws, r, 2)).startswith("week"):
            info = build_day_columns(ws, r)
            if not info:
                r += 1
                continue
            nums = []
            date_row = info["date_row"]
            for (c, idx) in info["day_cols"]:
                nums.append((idx, daynum(ws, date_row, c)))
            nums.sort()
            seq = [n for _, n in nums]
            dates = reconstruct_week(seq, month)
            if not dates:
                r += 1
                continue
            status = resolve_status(cell(ws, r, 3)) or "published"
            week_start = dates[0]
            # Walk shift rows until the next WEEK header.
            rr = r + 3
            current_dept = None
            while rr <= max_r and not norm(cell(ws, rr, 2)).startswith("week"):
                label = cell(ws, rr, 2)
                dept, role, current_dept = classify2(label, current_dept)
                if role and norm(role) in ROLE_FIX:
                    role = ROLE_FIX[norm(role)]
                if role:
                    for (c, day_idx, sc) in info["cols"]:
                        name = cell(ws, rr, c)
                        if not name:
                            continue
                        for frag in re.split(r"[\n/]", name):
                            frag = frag.strip()
                            if len(norm(frag)) < 2:
                                continue
                            pid = matcher.resolve(frag)
                            if not pid:
                                exceptions.append(
                                    (ws.title, week_start.isoformat(),
                                     label, frag, "unmatched"))
                                continue
                            wd = dates[day_idx]
                            if wd >= CLEAR_FROM:
                                continue
                            # Faithful to the grid: keep the row's role. Only
                            # when the row's role has no line do we fall back to
                            # the department's DVM (rescues a doctor whose name
                            # sits in a support row with no DVM-able template).
                            tgt_dept, tgt_role = dept, role
                            tpls = by_dr.get((tgt_dept or "", norm(tgt_role)))
                            if not tpls and pid in dvm_ids:
                                dd = dept or current_dept
                                if by_dr.get((dd or "", "dvm")):
                                    tgt_dept, tgt_role = dd, "DVM"
                                    tpls = by_dr.get((dd or "", "dvm"))
                            if not tpls:
                                exceptions.append(
                                    (ws.title, wd.isoformat(), label, frag,
                                     f"no-line({tgt_dept}/{tgt_role})"))
                                continue
                            loc_id = loc_ids.get(sc)
                            if not loc_id:
                                continue
                            placements.append({
                                "week_start": week_start, "status": status,
                                "tpl": tpls[0], "loc_id": loc_id,
                                "day": day_idx, "work_date": wd, "pid": pid,
                            })
                            dominant_role[pid][(tgt_dept or "", norm(tgt_role))] += 1
                            dominant_loc[pid][sc] += 1
                rr += 1
            r = rr
            continue
        r += 1


# ---------------------------------------------------------------------------
# CALLOUT 26 parsing
# ---------------------------------------------------------------------------

def parse_callout(ws, matcher, exceptions):
    """Return list of dicts: pid, work_date, status, note."""
    # Column -> date via the m/d labels in row 3.
    col_date = {}
    for c in range(13, ws.max_column + 1):
        v = cell(ws, 3, c)
        m = re.match(r"(\d{1,2})/(\d{1,2})", v)
        if m:
            try:
                col_date[c] = date(YEAR, int(m.group(1)), int(m.group(2)))
            except ValueError:
                pass
    out = []
    for r in range(4, ws.max_row + 1):
        name = cell(ws, r, 1)
        if not name:
            continue
        pid = matcher.resolve(name)
        if not pid:
            exceptions.append(("CALLOUT", "", name, name, "unmatched-employee"))
            continue
        for c, d in col_date.items():
            mark = cell(ws, r, c).upper()
            if mark not in ATT_MAP:
                continue
            if d >= CLEAR_FROM:
                continue
            font = ws.cell(row=r, column=c).font
            rgb = None
            try:
                rgb = font.color.rgb if font and font.color else None
            except Exception:
                rgb = None
            reported = not (isinstance(rgb, str) and rgb.upper().endswith("0000FF"))
            note = ("reported to Angie" if reported else "did not report to Angie")
            out.append({
                "pid": pid, "work_date": d, "status": ATT_MAP[mark],
                "note": f"2026 callout: {mark} ({note})",
            })
    return out


# ---------------------------------------------------------------------------
# VACATION CALENDAR 26 parsing
# ---------------------------------------------------------------------------

def parse_vacation(ws, matcher, exceptions):
    """Return dict pid -> sorted set of off-dates."""
    max_r, max_c = ws.max_row, ws.max_column
    # Locate month titles -> (row, start_col, month).
    titles = []
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = norm(cell(ws, r, c))
            if v in MONTH_NAMES:
                titles.append((r, c, MONTH_NAMES[v]))
    off = defaultdict(set)
    # Group titles by their row -> a horizontal band-set.
    title_rows = sorted({r for r, _, _ in titles})
    for ti, trow in enumerate(title_rows):
        row_titles = [(c, m) for (r, c, m) in titles if r == trow]
        row_titles.sort()
        next_title_row = title_rows[ti + 1] if ti + 1 < len(title_rows) else max_r + 1
        # Day-name header rows within this band-set.
        band_rows = [
            r for r in range(trow + 1, next_title_row)
            if any(norm(cell(ws, r, c)) in DAY_NAMES for c in range(1, max_c + 1))
        ]
        for bi, brow in enumerate(band_rows):
            date_row = brow + 1
            name_start = brow + 2
            name_end = (band_rows[bi + 1] if bi + 1 < len(band_rows)
                        else next_title_row) - 1
            for (start_col, month) in row_titles:
                # Map the 7 columns start_col..start_col+6 to real dates.
                nums = [daynum(ws, date_row, start_col + j) for j in range(7)]
                dates = reconstruct_week(nums, month)
                if not dates:
                    continue
                for rr in range(name_start, name_end + 1):
                    for j in range(7):
                        nm = cell(ws, rr, start_col + j)
                        if not nm:
                            continue
                        for frag in re.split(r"[\n/,]", nm):
                            frag = frag.strip()
                            if len(norm(frag)) < 2:
                                continue
                            pid = matcher.resolve(frag)
                            if not pid:
                                exceptions.append(
                                    ("VACATION", dates[j].isoformat(),
                                     frag, frag, "unmatched"))
                                continue
                            off[pid].add(dates[j])
    return {pid: sorted(ds) for pid, ds in off.items()}


def collapse_ranges(dates):
    """Sorted date list -> list of (start, end) consecutive runs."""
    runs = []
    for d in dates:
        if runs and d == runs[-1][1] + timedelta(days=1):
            runs[-1][1] = d
        else:
            runs.append([d, d])
    return [(a, b) for a, b in runs]


# ---------------------------------------------------------------------------
# SQL emission
# ---------------------------------------------------------------------------

def week_ensure_block(ws_date, status="published"):
    ws = ws_date.isoformat()
    title = ws_date.strftime("Week of %b %-d, %Y")
    return (
        "do $$ declare wid uuid; begin\n"
        f"  insert into sched_week (week_start, status, title) values "
        f"({sql_str(ws)}, {sql_str(status)}, {sql_str(title)}) "
        "on conflict (week_start) do nothing;\n"
        f"  select id into wid from sched_week where week_start = {sql_str(ws)};\n"
        "  if not exists (select 1 from sched_week_line where week_id = wid) then\n"
        "    insert into sched_week_line (week_id, template_id, department_id, role_id, label, start_time, end_time, sort_order)\n"
        "      select wid, id, department_id, role_id, label, start_time, end_time, sort_order from sched_shift_template where is_active;\n"
        "  end if;\n"
        "  if not exists (select 1 from sched_week_location where week_id = wid) then\n"
        "    insert into sched_week_location (week_id, location_id, sort_order)\n"
        "      select wid, id, sort_order from location where is_active;\n"
        "  end if;\n"
        "end $$;"
    )


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    matcher = Matcher()
    by_dr = load_templates()
    loc_ids = load_locations()
    dvm_ids = load_dvm_ids()

    exceptions = []
    placements = []
    dominant_role = defaultdict(Counter)
    dominant_loc = defaultdict(Counter)

    # 1) Month schedule tabs.
    for tab, month in MONTH_TABS.items():
        if tab in wb.sheetnames:
            parse_month_tab(wb[tab], month, matcher, by_dr, dvm_ids, loc_ids,
                            placements, dominant_role, dominant_loc, exceptions)

    # 2) CALLOUT attendance.
    callouts = parse_callout(wb["CALLOUT 26"], matcher, exceptions)

    # 3) VACATION PTO.
    pto = parse_vacation(wb["VACATION CALENDAR 26"], matcher, exceptions)

    # --- Determine every week we touch (schedule + callout) -----------------
    week_status = {}
    for p in placements:
        week_status.setdefault(p["week_start"], p["status"])
    for c in callouts:
        week_status.setdefault(sunday_of(c["work_date"]), "published")

    # --- Emit SQL -----------------------------------------------------------
    out = [
        "-- Green Dog Ops — Schedule Upload workbook import.",
        "-- Generated by scripts/import_schedule_upload.py. Idempotent overlay:",
        "-- weeks/lines created if missing; assignments never wiped/duplicated.",
        "set search_path = greendogops, public;",
        "",
        "-- ============ 1. Ensure weeks (with line + location snapshots) ======",
    ]
    for ws_date in sorted(week_status):
        out.append(week_ensure_block(ws_date, week_status[ws_date]))
    out.append("")
    out.append("-- ============ 2. Schedule assignments (overlay) ============")
    for p in placements:
        ws = p["week_start"].isoformat()
        out.append(
            "do $$ declare wid uuid; begin\n"
            f"  select id into wid from sched_week where week_start = {sql_str(ws)};\n"
            "  insert into sched_assignment (week_id, line_id, location_id, person_id, day_of_week, work_date, attendance_status)\n"
            f"    select wid, wl.id, {sql_str(p['loc_id'])}, {sql_str(p['pid'])}, {p['day']}, {sql_str(p['work_date'].isoformat())}, 'scheduled'\n"
            f"    from sched_week_line wl where wl.week_id = wid and wl.template_id = {sql_str(p['tpl'])}\n"
            "      and not exists (select 1 from sched_assignment a where a.week_id = wid\n"
            f"        and a.line_id = wl.id and a.location_id = {sql_str(p['loc_id'])}\n"
            f"        and a.day_of_week = {p['day']} and a.person_id = {sql_str(p['pid'])})\n"
            "    limit 1;\n"
            "end $$;"
        )

    # --- 3. PTO days + approved time-off ranges + flip scheduled -> pto ------
    out.append("")
    out.append("-- ============ 3. PTO (itemized days + approved ranges) ======")
    for pid, dates in pto.items():
        for d in dates:
            ds = d.isoformat()
            out.append(
                "insert into person_pto_day (person_id, pto_date, note)\n"
                f"  select {sql_str(pid)}, {sql_str(ds)}, {sql_str('Vacation calendar 2026')}\n"
                "  where not exists (select 1 from person_pto_day where "
                f"person_id = {sql_str(pid)} and pto_date = {sql_str(ds)});"
            )
            out.append(
                "update sched_assignment set attendance_status = 'pto'\n"
                f"  where person_id = {sql_str(pid)} and work_date = {sql_str(ds)} "
                "and attendance_status = 'scheduled';"
            )
        for a, b in collapse_ranges(dates):
            out.append(
                "insert into person_time_off (person_id, kind, status, start_date, end_date, note)\n"
                f"  select {sql_str(pid)}, 'pto', 'approved', {sql_str(a.isoformat())}, {sql_str(b.isoformat())}, {sql_str('Vacation calendar 2026')}\n"
                "  where not exists (select 1 from person_time_off where "
                f"person_id = {sql_str(pid)} and start_date = {sql_str(a.isoformat())} "
                f"and end_date = {sql_str(b.isoformat())} and kind = 'pto');"
            )

    # --- 4. CALLOUT attendance (update existing; else create minimal) -------
    out.append("")
    out.append("-- ============ 4. Attendance markings (CALLOUT 26) ==========")
    for c in callouts:
        pid = c["pid"]
        wd = c["work_date"]
        ws = sunday_of(wd).isoformat()
        dom = dominant_role[pid].most_common(1)
        dom_tpl = by_dr.get(dom[0][0], [None])[0] if dom else None
        dloc = dominant_loc[pid].most_common(1)
        loc_sc = dloc[0][0] if dloc else "SO"
        loc_id = loc_ids.get(loc_sc) or loc_ids.get("SO")
        note = c["note"]
        out.append(
            "do $$ declare wid uuid; lid uuid; begin\n"
            f"  select id into wid from sched_week where week_start = {sql_str(ws)};\n"
            f"  if exists (select 1 from sched_assignment where person_id = {sql_str(pid)} and work_date = {sql_str(wd.isoformat())}) then\n"
            f"    update sched_assignment set attendance_status = {sql_str(c['status'])},\n"
            f"      attendance_note = {sql_str(note)}\n"
            f"      where person_id = {sql_str(pid)} and work_date = {sql_str(wd.isoformat())}\n"
            "        and attendance_status in ('scheduled','present');\n"
            "  else\n"
            + (f"    select wl.id into lid from sched_week_line wl where wl.week_id = wid and wl.template_id = {sql_str(dom_tpl)} limit 1;\n"
               if dom_tpl else "")
            + "    if lid is null then select id into lid from sched_week_line where week_id = wid order by sort_order limit 1; end if;\n"
            "    if lid is not null then\n"
            "      insert into sched_assignment (week_id, line_id, location_id, person_id, day_of_week, work_date, attendance_status, attendance_note, added_post_publish)\n"
            f"        values (wid, lid, {sql_str(loc_id)}, {sql_str(pid)}, {sheet_dow(wd)}, {sql_str(wd.isoformat())}, {sql_str(c['status'])}, {sql_str(note + ' (no scheduled shift)')}, true);\n"
            "    end if;\n"
            "  end if;\n"
            "end $$;"
        )

    Path(".data").mkdir(exist_ok=True)
    Path(OUT_SQL).write_text("\n".join(out) + "\n")

    # Also emit statement-aligned chunks (the Management API runs the whole body
    # as one request; 8 MB is too big, so apply in idempotent slices).
    chunk_dir = Path(".data/schedule_upload_chunks")
    chunk_dir.mkdir(exist_ok=True)
    for old in chunk_dir.glob("part_*.sql"):
        old.unlink()
    stmts = [s for s in out if s.strip() and not s.lstrip().startswith("--")
             and s.strip() != "set search_path = greendogops, public;"]
    per = 500
    n_chunks = 0
    for i in range(0, len(stmts), per):
        body = "set search_path = greendogops, public;\n" + \
            "\n".join(stmts[i:i + per]) + "\n"
        (chunk_dir / f"part_{i // per:03d}.sql").write_text(body)
        n_chunks += 1

    # --- Report -------------------------------------------------------------
    by_reason = Counter(r for *_, r in ((e[0], e[4]) for e in exceptions))
    rep = []
    rep.append("=" * 72)
    rep.append("SCHEDULE UPLOAD IMPORT — dry parse summary")
    rep.append("=" * 72)
    rep.append(f"Weeks touched:          {len(week_status)}")
    rep.append(f"Schedule placements:    {len(placements)}")
    rep.append(f"Callout markings:       {len(callouts)}")
    rep.append(f"PTO people / days:      {len(pto)} / {sum(len(v) for v in pto.values())}")
    rep.append(f"Exceptions:             {len(exceptions)}")
    rep.append("")
    rep.append("Exceptions by reason:")
    for reason, n in sorted(by_reason.items(), key=lambda x: -x[1]):
        rep.append(f"  {n:5}  {reason}")
    rep.append("")
    rep.append(f"Unmatched names ({len(matcher.unmatched)} distinct):")
    for nm, n in matcher.unmatched.most_common():
        rep.append(f"  {n:4}  {nm}")
    report = "\n".join(rep)
    Path(OUT_REPORT).write_text(report + "\n")
    print(report)
    print(f"\nSQL   -> {OUT_SQL}")
    print(f"Report-> {OUT_REPORT}")
    print(f"Chunks-> {chunk_dir}/ ({n_chunks} files)")


if __name__ == "__main__":
    main()
