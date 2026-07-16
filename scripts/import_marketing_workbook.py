#!/usr/bin/env python3
"""
Parse public/Marketing Spreadsheets (2).xlsx into SQL for the marketing module.

Loads the high-value, in-scope sheets into the existing marketing tables:
  - 2026 Event Recaps        -> greendogops.marketing_event (completed, w/ ROI)
  - DRE/JENN MARKETING SPEND  -> greendogops.marketing_budget_entry (admin-only)
  - Adopta 2026 Budget        -> greendogops.marketing_budget_entry
  - 2026 Event HostsSources   -> greendogops.marketing_resource (event sources)
  - RESOURCES & PW            -> greendogops.marketing_resource (tools; NO
                                 passwords, NO credential-doc links)

Emits idempotent INSERT ... WHERE NOT EXISTS statements to stdout. Apply with:
    python3 scripts/import_marketing_workbook.py | scripts/supabase-sql.sh
"""
import datetime
import re
import sys

import openpyxl

XLSX = "public/Marketing Spreadsheets (2).xlsx"


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def as_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = clean(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.datetime.strptime(s.split()[0], fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def as_num(v):
    s = clean(v)
    if not s:
        return None
    if s.lower() == "free":
        return 0.0
    t = s.replace("$", "").replace(",", "").strip()
    try:
        return float(t)
    except ValueError:
        return None


def as_int(v):
    n = as_num(v)
    if n is None:
        return None
    return int(n) if float(n).is_integer() else None


def q(v):
    """SQL string literal."""
    if v is None:
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def qn(v):
    return "NULL" if v is None else repr(float(v))


def qi(v):
    return "NULL" if v is None else str(int(v))


wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
out = []
counts = {"event": 0, "budget": 0, "resource": 0}

# ---------------------------------------------------------------------------
# 1) Events — from "2026 Event Recaps"
# ---------------------------------------------------------------------------
ws = wb["2026 Event Recaps"]
for r in list(ws.iter_rows(values_only=True))[1:]:
    r = list(r) + [None] * (13 - len(r))
    name = clean(r[0])
    if not name:
        continue
    # Skip the year-separator rows (e.g. "2026", "2025").
    if re.fullmatch(r"\d{4}(\.0)?", name):
        continue
    starts_on = as_date(r[1])
    cost = as_num(r[4])
    attendees = as_int(r[5])
    signups = as_int(r[6])
    appointments = as_int(r[7])
    products_sold = clean(r[8])
    redemption_codes = clean(r[9])
    coupons = as_int(r[10])
    client_spend = as_num(r[11])
    feedback_parts = [clean(r[12])]
    # Preserve non-numeric raw values that don't fit the integer columns.
    if attendees is None and clean(r[5]):
        feedback_parts.append(f"Attendees: {clean(r[5])}")
    if signups is None and clean(r[6]):
        feedback_parts.append(f"Sign-ups: {clean(r[6])}")
    if cost is None and clean(r[4]):
        feedback_parts.append(f"Cost: {clean(r[4])}")
    feedback = " | ".join(p for p in feedback_parts if p) or None
    out.append(
        f"insert into greendogops.marketing_event "
        f"(name, event_type, status, starts_on, location, clinic_served, cost, "
        f"attendees, signups, appointments, products_sold, redemption_codes, "
        f"coupons_redeemed, client_spend, feedback) "
        f"select {q(name)}, 'third_party', 'completed', "
        f"{('date ' + q(starts_on)) if starts_on else 'NULL'}, {q(clean(r[2]))}, "
        f"{q(clean(r[3]))}, {qn(cost)}, {qi(attendees)}, {qi(signups)}, "
        f"{qi(appointments)}, {q(products_sold)}, {q(redemption_codes)}, "
        f"{qi(coupons)}, {qn(client_spend)}, {q(feedback)} "
        f"where not exists (select 1 from greendogops.marketing_event e "
        f"where e.name = {q(name)} and e.starts_on is not distinct from "
        f"{('date ' + q(starts_on)) if starts_on else 'NULL'});"
    )
    counts["event"] += 1

# ---------------------------------------------------------------------------
# 2) Budget entries — DRE / JENN spending + Adopta budget
# ---------------------------------------------------------------------------
def emit_budget(entry_date, business, description, amount, paid_by,
                payment_method, category, status, receipt, notes):
    if amount is None:
        return
    out.append(
        f"insert into greendogops.marketing_budget_entry "
        f"(entry_date, category, business, description, amount, paid_by, "
        f"payment_method, status, receipt_submitted, notes) "
        f"select {('date ' + q(entry_date)) if entry_date else 'current_date'}, "
        f"{q(category)}, {q(business)}, {q(description)}, {qn(amount)}, "
        f"{q(paid_by)}, {q(payment_method)}, {q(status)}, "
        f"{'true' if receipt else 'false'}, {q(notes)} "
        f"where not exists (select 1 from greendogops.marketing_budget_entry b "
        f"where b.business is not distinct from {q(business)} "
        f"and b.entry_date is not distinct from "
        f"{('date ' + q(entry_date)) if entry_date else 'current_date'} "
        f"and b.amount = {qn(amount)});"
    )
    counts["budget"] += 1


for sheet in ("DRE MARKETING SPENDING", "JENN MARKETING SPENDING"):
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    section = ""
    for r in rows[3:]:
        r = list(r) + [None] * (11 - len(r))
        cells = [clean(c) for c in r]
        amount = as_num(r[5])
        business = clean(r[1])
        # Section marker: text present but no amount and no business.
        if amount is None and not business:
            label = next((c for c in cells if c), None)
            if label:
                section = label
            continue
        receipt = "reconcil" in section.lower()
        notes = " · ".join(
            p for p in [clean(r[0]), clean(r[8]), clean(r[10]), section] if p
        ) or None
        emit_budget(
            as_date(r[4]), business, clean(r[2]), amount, clean(r[6]),
            clean(r[7]), "Marketing", "paid", receipt, notes,
        )

ws = wb["Adopta 2026 Budget"]
for r in list(ws.iter_rows(values_only=True))[1:]:
    r = list(r) + [None] * (8 - len(r))
    amount = as_num(r[5])
    vendor = clean(r[1])
    if amount is None or not vendor:
        continue
    emit_budget(
        as_date(r[0]), vendor, clean(r[2]), amount, clean(r[6]),
        clean(r[7]), "Adoptapalooza", "paid", False, None,
    )

# ---------------------------------------------------------------------------
# 3) Resources — event sources + real tools (NO passwords / credential docs)
# ---------------------------------------------------------------------------
def emit_resource(name, category, url, description, credential_note):
    if not name:
        return
    out.append(
        f"insert into greendogops.marketing_resource "
        f"(name, category, url, description, credential_note) "
        f"select {q(name)}, {q(category)}, {q(url)}, {q(description)}, "
        f"{q(credential_note)} "
        f"where not exists (select 1 from greendogops.marketing_resource m "
        f"where lower(m.name) = lower({q(name)}));"
    )
    counts["resource"] += 1


# Event sources: "2026 Event HostsSources" (cols: name, Jan..Dec, cost, notes)
ws = wb["2026 Event HostsSources"]
for r in list(ws.iter_rows(values_only=True))[2:]:
    r = list(r) + [None] * (15 - len(r))
    label = clean(r[0])
    if not label:
        continue
    url = label if label.lower().startswith("http") else None
    name = label
    if url:
        # Derive a friendly name from the domain.
        name = re.sub(r"^https?://(www\.)?", "", label).split("/")[0]
    cost = clean(r[13])
    notes = clean(r[14])
    desc = " — ".join(p for p in [notes, cost] if p) or "Event-sourcing calendar / listing."
    emit_resource(name, "membership", url, desc, None)

# Tools: "RESOURCES & PW" — import only real product tools with a non-Google
# URL; skip password-vault doc links and any password data entirely.
ws = wb["RESOURCES & PW"]
for r in list(ws.iter_rows(values_only=True)):
    r = list(r) + [None] * (8 - len(r))
    name = clean(r[0])
    url = clean(r[1])
    if not name or not url or not url.lower().startswith("http"):
        continue
    low = (name + " " + url).lower()
    # Never surface password/credential documents.
    if "password" in low or "docs.google.com" in low or "drive.google.com" in low:
        continue
    emit_resource(name, "tool", url, None, "Login in the marketing credentials vault.")

sys.stderr.write(
    f"Generated: {counts['event']} events, {counts['budget']} budget entries, "
    f"{counts['resource']} resources.\n"
)
print("\n".join(out))
