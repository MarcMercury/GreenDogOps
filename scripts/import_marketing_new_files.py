#!/usr/bin/env python3
"""
Parse the newer marketing workbooks into SQL for the marketing module:
  - "Current Promotions / Groupon List.xlsx"
      Current Promotions / Upcoming / Expired / Influencers / Gift Certs /
      Vetstoria Widgets  -> greendogops.marketing_promotion
  - "Green Dog / GeniusVets Project Management.xlsx"
      GD Project Management Tracker  -> greendogops.marketing_initiative (GeniusVets)
      AETNA EVENTS RSVP TRACKER      -> greendogops.marketing_event (completed)
      DVM Open House Check Ins       -> attendees of "Industry Open House"
      Grand Opening Check Ins        -> attendees of "Green Dog Land"

Emits idempotent SQL to stdout. Apply with:
    python3 scripts/import_marketing_new_files.py | scripts/supabase-sql.sh
"""
import datetime
import re
import sys

import openpyxl

PROMO_XLSX = "public/Current Promotions _ Groupon List.xlsx"
GV_XLSX = "public/Green Dog _ GeniusVets Project Management .xlsx"


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def as_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = clean(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.datetime.strptime(s.split()[0], fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return None


def as_num(v):
    s = clean(v)
    if not s:
        return None
    t = s.replace("$", "").replace(",", "").replace("%", "").strip()
    try:
        return float(t)
    except ValueError:
        return None


def as_int(v):
    n = as_num(v)
    return int(n) if n is not None and float(n).is_integer() else None


def q(v):
    if v is None:
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def qn(v):
    return "NULL" if v is None else repr(float(v))


def qi(v):
    return "NULL" if v is None else str(int(v))


out = []
counts = {"promo": 0, "event": 0, "attendee": 0, "initiative": 0}


def looks_header(cells):
    joined = " ".join((c or "") for c in cells).upper()
    return "PROMOTION" in joined and "PLACEMENT" in joined


# ---------------------------------------------------------------------------
# Promotions
# ---------------------------------------------------------------------------
def emit_promo(name, placement, status, ptype, duration, discount_text,
               discount_amount, code, line_item, redeem, url, booking_url,
               rules, appts, notes):
    if not name:
        return
    out.append(
        "insert into greendogops.marketing_promotion "
        "(name, placement, status, promo_type, duration_text, discount_text, "
        "discount_amount, product_code, ezyvet_line_item, how_to_redeem, "
        "promo_url, booking_url, rules, appointments, notes) "
        f"select {q(name)}, {q(placement)}, {q(status)}, {q(ptype)}, "
        f"{q(duration)}, {q(discount_text)}, {qn(discount_amount)}, {q(code)}, "
        f"{q(line_item)}, {q(redeem)}, {q(url)}, {q(booking_url)}, {q(rules)}, "
        f"{qi(appts)}, {q(notes)} "
        "where not exists (select 1 from greendogops.marketing_promotion p "
        f"where p.name = {q(name)} and p.status = {q(status)});"
    )
    counts["promo"] += 1


wb = openpyxl.load_workbook(PROMO_XLSX, read_only=True, data_only=True)


def promo_sheet(title, status, ptype, has_code=True):
    ws = wb[title]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows:
        r = list(r) + [None] * 10
        cells = [clean(c) for c in r]
        name = cells[0]
        if not name or looks_header(cells) or name.upper().startswith(("CURRENT PROMOTIONS", "NEW UPCOMING", "EXPIRED PROMOTION", "THESE ARE", "EXAMPLE")):
            continue
        if has_code:
            emit_promo(name, cells[1], status, ptype, cells[2], cells[3],
                       as_num(cells[3]), cells[4], cells[5], cells[6], cells[7],
                       None, cells[8], None, None)
        else:  # Influencers sheet has no product-code column
            emit_promo(name, cells[1], status, ptype, cells[2], cells[3],
                       as_num(cells[3]), None, cells[4], cells[5], cells[6],
                       None, cells[7], None, None)


def find_sheet(prefix):
    for ws in wb.worksheets:
        if ws.title.strip().upper().startswith(prefix.upper()):
            return ws.title
    return None


for prefix, status, ptype, has_code in [
    ("Current Promotions", "active", "standard", True),
    ("EXPIRED COUPONS", "expired", "standard", True),
    ("DRE & GLADYS", "upcoming", "standard", True),
    ("Gift Certificates", "active", "gift_certificate", True),
    ("Influencers", "active", "influencer", False),
]:
    t = find_sheet(prefix)
    if t:
        promo_sheet(t, status, ptype, has_code)

# Vetstoria widgets (different shape: NOTES | Name | Link | Widget Link | Webpage | Appointments)
tw = find_sheet("Vetstoria")
if tw:
    for r in list(wb[tw].iter_rows(values_only=True))[1:]:
        r = list(r) + [None] * 7
        notes, _num, link_name, widget_link, webpage, appts = (
            clean(r[0]), r[1], clean(r[2]), clean(r[3]), clean(r[4]), as_int(r[6]),
        )
        if not link_name:
            continue
        status = "active"
        wl = (webpage or "").lower()
        if "expired" in wl or "not found" in wl or "unpublish" in wl:
            status = "expired"
        emit_promo(link_name, "Vetstoria booking widget", status, "widget", None,
                   None, None, None, None, None, None, widget_link, None, appts, notes)

# ---------------------------------------------------------------------------
# GeniusVets workbook: initiatives, events, attendees
# ---------------------------------------------------------------------------
gv = openpyxl.load_workbook(GV_XLSX, read_only=True, data_only=True)


def gv_sheet(prefix):
    for ws in gv.worksheets:
        if ws.title.strip().upper().startswith(prefix.upper()):
            return ws
    return None


def map_status(raw):
    r = (raw or "").lower()
    if "approv" in r and "need" not in r:
        return "done"
    if "complete" in r or "done" in r:
        return "done"
    if "need" in r or "hold" in r:
        return "blocked"
    if "progress" in r:
        return "in_progress"
    return "planned"


# GD Project Management Tracker -> initiatives
ws = gv_sheet("GD Project Management")
if ws:
    for r in list(ws.iter_rows(values_only=True))[1:]:
        r = list(r) + [None] * 10
        task = clean(r[0])
        if not task or task.upper().startswith("TASK"):
            continue
        owner = clean(r[3]) or clean(r[2])
        status = map_status(clean(r[4]))
        due = as_date(r[5])
        deliverable = clean(r[6])
        tmpl = clean(r[7])
        notes = " · ".join(p for p in [clean(r[8]), clean(r[9])] if p) or None
        links = "'[]'::jsonb"
        if tmpl and tmpl.lower().startswith("http"):
            links = f"jsonb_build_array(jsonb_build_object('label','Template','url',{q(tmpl)}))"
        out.append(
            "insert into greendogops.marketing_initiative "
            "(title, category, status, priority, owner_name, partner_name, "
            "next_action, notes, links) "
            f"select {q(task[:200])}, 'pr', {q(status)}, 'medium', {q(owner)}, "
            f"'GeniusVets', {q(deliverable)}, {q(notes)}, {links} "
            "where not exists (select 1 from greendogops.marketing_initiative i "
            f"where i.title = {q(task[:200])} and i.partner_name = 'GeniusVets');"
        )
        counts["initiative"] += 1

# AETNA EVENTS RSVP TRACKER -> events (skip generic "CE Event")
KNOWN_EVENT_DATES = {
    "Green Dog Land": "2025-03-29",
    "Industry Open House": "2025-03-22",
}
ws = gv_sheet("AETNA EVENTS RSVP")
if ws:
    for r in list(ws.iter_rows(values_only=True))[1:]:
        r = list(r) + [None] * 8
        name = clean(r[0])
        if not name or name.lower() == "ce event":
            continue
        rsvp_total = as_int(r[1])
        checkins = as_int(r[5])
        page_visits = as_int(r[6])
        starts = KNOWN_EVENT_DATES.get(name)
        feedback = f"RSVPs: {rsvp_total or '—'} · Check-ins: {checkins or '—'} · Page visits: {page_visits or '—'}"
        out.append(
            "insert into greendogops.marketing_event "
            "(name, event_type, status, starts_on, location, owner_name, "
            "attendees, signups, feedback) "
            f"select {q(name)}, 'hosted', 'completed', "
            f"{('date ' + q(starts)) if starts else 'NULL'}, 'Van Nuys', "
            f"'Marketing', {qi(checkins)}, {qi(rsvp_total)}, {q(feedback)} "
            "where not exists (select 1 from greendogops.marketing_event e "
            f"where e.name = {q(name)});"
        )
        counts["event"] += 1


def emit_attendee(event_name, name, email, phone, atype, is_new, notes):
    if not (name or email or phone):
        return
    out.append(
        "insert into greendogops.marketing_event_attendee "
        "(event_id, name, email, phone, attendee_type, is_new_client, notes) "
        f"select e.id, {q(name)}, {q(email)}, {q(phone)}, {q(atype)}, "
        f"{'true' if is_new else 'false'}, {q(notes)} "
        f"from greendogops.marketing_event e where e.name = {q(event_name)} "
        "and not exists (select 1 from greendogops.marketing_event_attendee a "
        f"where a.event_id = e.id and a.name is not distinct from {q(name)} "
        f"and a.email is not distinct from {q(email)});"
    )
    counts["attendee"] += 1


# DVM Open House Check Ins -> attendees of Industry Open House
ws = gv_sheet("DVM Open House Check Ins")
if ws:
    for r in list(ws.iter_rows(values_only=True))[1:]:
        r = list(r) + [None] * 8
        name = clean(r[2])
        if not name or name.upper() == "NAME":
            continue
        practice = clean(r[6])
        heard = clean(r[7])
        notes = " · ".join(p for p in [practice, heard] if p) or None
        emit_attendee("Industry Open House", name, clean(r[4]), clean(r[5]), "lead", False, notes)

# Grand Opening Check Ins -> attendees of Green Dog Land
ws = gv_sheet("Grand Opening Check Ins")
if ws:
    for r in list(ws.iter_rows(values_only=True))[2:]:
        r = list(r) + [None] * 13
        name = clean(r[3])
        if not name or name.upper() == "NAME":
            continue
        email = clean(r[4])
        phone = clean(r[5])
        heard = clean(r[6])
        current_client = (clean(r[8]) or "").lower()
        is_new = current_client.startswith("no")
        atype = "returning" if current_client.startswith("yes") else "new_client"
        emit_attendee("Green Dog Land", name, email, phone, atype, is_new, heard)

sys.stderr.write(
    f"Generated: {counts['promo']} promotions, {counts['event']} events, "
    f"{counts['initiative']} initiatives, {counts['attendee']} attendees.\n"
)
print("\n".join(out))
