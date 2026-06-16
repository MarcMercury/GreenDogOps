"""
Import Green Dog roster (current + former employees) from the Excel workbook
into the greendogops schema. Emits SQL to stdout; pipe into scripts/supabase-sql.sh.

Usage:
    python3 scripts/import_roster.py > /tmp/roster_import.sql
    scripts/supabase-sql.sh -f /tmp/roster_import.sql
"""
import datetime
import openpyxl

PATH = "public/Untitled spreadsheet.xlsx"
wb = openpyxl.load_workbook(PATH, read_only=True, data_only=True)


def sql_str(v):
    if v is None or v == "":
        return "null"
    return "'" + str(v).replace("'", "''") + "'"


def sql_num(v):
    if v is None or v == "":
        return "null"
    try:
        f = float(v)
        return repr(f)
    except (TypeError, ValueError):
        return "null"


def sql_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return "'" + v.strftime("%Y-%m-%d") + "'"
    return "null"


def sql_bool(v):
    if v is None or v == "":
        return "null"
    s = str(v).strip().lower()
    if s in ("true", "yes", "y", "1", "done", "completed", "complete"):
        return "true"
    if s in ("false", "no", "n", "0"):
        return "false"
    return "null"  # 'N/A' etc.


def work_loc(v):
    if not v:
        return "null"
    s = str(v).strip().lower()
    if "remote" in s:
        return "'remote'"
    if "hybrid" in s:
        return "'hybrid'"
    if "house" in s or "in-house" in s or "in house" in s:
        return "'in_house'"
    return "null"


def flsa(v):
    if not v:
        return "null"
    s = str(v).strip().lower()
    if "non" in s:
        return "'non_exempt'"
    if "exempt" in s:
        return "'exempt'"
    return "null"


def schedule(v):
    if not v:
        return "null"
    s = str(v).strip().lower()
    if "full" in s:
        return "'full_time'"
    if "part" in s:
        return "'part_time'"
    if "diem" in s or "per diem" in s:
        return "'per_diem'"
    if "contract" in s:
        return "'contractor'"
    return "null"


def days_per_week(v):
    # roster column sometimes holds a stray date/year; only accept 0-14 days
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        f = float(v)
        if 0 <= f <= 14:
            return repr(f)
    return "null"


print("begin;")

# ---- CURRENT EMPLOYEES (roster sheet, header on row index 1) ----
ws = wb["roster"]
rows = list(ws.iter_rows(values_only=True))
for r in rows[2:]:
    if not r or not r[0] or str(r[0]).strip() == "":
        continue
    full_name = r[0]
    # Skip obvious non-person section rows
    if str(full_name).strip().upper() == full_name and len(str(full_name)) > 40:
        continue
    compliance = {
        "offer_letter_completed": r[26],
        "handbook_signed": r[27],
        "onboarding_completed": r[28],
        "benefits_completed": r[29],
        "sexual_harassment_training_date": str(r[30]) if r[30] else None,
        "harassment_pay": r[31],
        "background_check_processed": r[32],
        "safety_training": r[33],
        "emergency_contact_form": r[34],
        "contract_sent": r[35],
        "contract_signed": r[36],
        "ce_contract_sent": r[38],
        "ce_contract_signed": r[39],
        "immigration_agreement_sent": r[40],
        "immigration_agreement_signed": r[41],
        "licenses_tracked": r[42],
    }
    import json
    comp_json = json.dumps({k: (str(v) if v is not None else None) for k, v in compliance.items()})
    print("with p as (")
    print(f"  insert into greendogops.person (status, first_name, last_name, grid_name, full_name, date_of_birth, postal_code, work_location_type, is_active)")
    print(f"  values ('employee', {sql_str(r[2])}, {sql_str(r[3])}, {sql_str(r[1])}, {sql_str(r[0])}, {sql_date(r[4])}, {sql_str(int(r[5]) if isinstance(r[5], float) else r[5])}, {work_loc(r[6])}, true)")
    print("  returning id)")
    print("insert into greendogops.person_employment (person_id, offer_title, adp_job_title, flsa_status, work_schedule, days_per_week, hire_date, latest_wage_change_date, current_rate, previous_rate, biweekly_wage, annual_wages, pto_allotment, pto_used, pto_available, pto_notes, ce_budget, ce_used, ce_remaining, compliance)")
    print(f"select id, {sql_str(r[7])}, {sql_str(r[8])}, {flsa(r[10])}, {schedule(r[11])}, {days_per_week(r[12])}, {sql_date(r[9])}, {sql_date(r[13])}, {sql_num(r[14])}, {sql_num(r[15])}, {sql_num(r[16])}, {sql_num(r[17])}, {sql_str(r[18])}, {sql_num(r[20])}, {sql_num(r[21])}, {sql_str(r[22])}, {sql_num(r[23])}, {sql_num(r[24])}, {sql_num(r[25])}, '{comp_json}'::jsonb from p;")

# ---- FORMER EMPLOYEES (former employees sheet, header on row index 0) ----
ws = wb["former employees"]
rows = list(ws.iter_rows(values_only=True))
for r in rows[2:]:
    if not r or not r[0] or str(r[0]).strip() == "":
        continue
    full_name = str(r[0]).strip()
    if full_name.upper() == full_name and len(full_name) > 40:
        continue
    sep_type = "null"
    qf = (str(r[7]).strip().lower() if r[7] else "")
    if "quit" in qf:
        sep_type = "'quit'"
    elif "fired" in qf:
        sep_type = "'fired'"
    elif "laid" in qf:
        sep_type = "'laid_off'"
    print("with p as (")
    print(f"  insert into greendogops.person (status, first_name, last_name, full_name, date_of_birth, postal_code, work_location_type, is_active)")
    print(f"  values ('former', {sql_str(r[1])}, {sql_str(r[2])}, {sql_str(r[0])}, {sql_date(r[3])}, {sql_str(int(r[4]) if isinstance(r[4], float) else r[4])}, {work_loc(r[5])}, false)")
    print("  returning id)")
    print("insert into greendogops.person_employment (person_id, adp_job_title, flsa_status, work_schedule, days_per_week, hire_date, last_review_date, current_rate, latest_wage_change_date, biweekly_wage, annual_wages, pto_used, pto_available, separation_date, separation_type, separation_letter_signed)")
    print(f"select id, {sql_str(r[9])}, {flsa(r[11])}, {schedule(r[12])}, {days_per_week(r[13])}, {sql_date(r[10])}, {sql_date(r[14])}, {sql_num(r[15])}, {sql_date(r[17])}, {sql_num(r[19])}, {sql_num(r[20])}, {sql_num(r[23])}, {sql_num(r[24])}, {sql_date(r[6])}, {sep_type}, {sql_bool(r[8])} from p;")

print("commit;")
wb.close()
