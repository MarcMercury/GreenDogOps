#!/usr/bin/env python3
"""Parse the veterinary ASSOCIATIONS / GROUPS from the CE Outreach workbook and
emit SQL inserts into greendogops.crm_organization as Vendor-CRM rows
(org_type='med_ops', subtype categorizes them). These are professional
associations, affinity groups, alumni groups, conferences, boards, equipment
vendors and industry media — NOT individual people and NOT referral clinics.

Two sheets are merged (keyed by acronym, then by significant-token core, then by
normalized name):
  'CE Research AssociationsGroupsE' : contact details (email/phone/director/site).
                                       Only the sections ABOVE "LOCAL CLINICS".
  'Local Veterinary Groups'         : acronyms + descriptive notes.

NON-DESTRUCTIVE / RE-RUNNABLE: each insert is guarded by NOT EXISTS on a
normalized name match so we never duplicate an org already in crm_organization.

Usage:
    python scripts/import_ce_associations.py | ./scripts/supabase-sql.sh
"""
import re
import sys

import openpyxl

SRC = ".data/CE Outreach Contact List_2026-06-17.xlsx"
SOURCE_TAG = "ce_associations_2026"

EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
URL_RE = re.compile(r"(https?://[^\s|]+|www\.[^\s|]+|[A-Za-z0-9.\-]+\.(?:com|org|net|edu|ph)(?:/[^\s|]*)?)")
PHONE_RE = re.compile(r"\(?\d{3}\)?[\s.\-]?\d{3}[\s.\-]?\d{4}")
ACRO_RE = re.compile(r"^[A-Z][A-Z0-9.\-]{1,7}$")

STOPWORDS = {
    "veterinary", "veterinarians", "veterinarian", "medical", "medicine",
    "association", "associations", "vma", "vmc", "group", "groups", "of", "the",
    "a", "an", "inc", "society", "college", "school", "network", "professionals",
    "american", "national", "and", "for", "in", "ca", "california", "southern",
    "based", "nonprofit", "supporting", "ce", "aid",
}


def cell(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\u00a0", " ")).strip()


def normname(name):
    return re.sub(r"[^a-z0-9]", "", name.lower())


def clean_name(name):
    """Drop a trailing descriptive parenthetical (overly long or unbalanced)."""
    name = cell(name)
    idx = name.find(" (")
    if idx != -1:
        head, tail = name[:idx].strip(), name[idx + 1:]
        if len(tail) > 40 or tail.count("(") != tail.count(")"):
            return head
    return name


def core_tokens(name):
    toks = re.findall(r"[a-z0-9]+", name.lower())
    sig = frozenset(t for t in toks if t not in STOPWORDS and len(t) > 2)
    return sig


def clean_acro(v):
    v = cell(v)
    if not v or v in ("-", "—"):
        return None
    if ACRO_RE.match(v) and not v.lower().endswith((".com", ".org", ".net")):
        return v.upper().rstrip(".")
    return None


def first_email(v):
    v = cell(v)
    if not v:
        return None, None
    if "need to fill out" in v.lower() or "need to add" in v.lower():
        return None, "Register via their website to obtain contact"
    m = EMAIL_RE.findall(v)
    if not m:
        return None, None
    primary = m[0].lower()
    extra = None
    if len(m) > 1:
        extra = "Alt emails: " + ", ".join(x.lower() for x in m[1:])
    return primary, extra


def first_phone(v):
    v = cell(v)
    if not v:
        return None, None
    m = PHONE_RE.findall(v)
    if not m:
        return None, None
    primary = m[0]
    extra = None
    leftover = v
    if "fax" in v.lower():
        extra = "Phone detail: " + v
    return primary, extra


def first_url(v):
    v = cell(v)
    if not v:
        return None
    m = URL_RE.search(v)
    if not m:
        return None
    url = m.group(0).rstrip(".,")
    return url


class Org:
    __slots__ = ("name", "acro", "subtype", "email", "phone", "contact_name",
                 "website", "notes", "core")

    def __init__(self):
        self.name = None
        self.acro = None
        self.subtype = None
        self.email = None
        self.phone = None
        self.contact_name = None
        self.website = None
        self.notes = []
        self.core = frozenset()


by_acro = {}
by_core = {}
by_name = {}
orgs = []


def find(name, acro):
    if acro and acro in by_acro:
        return by_acro[acro]
    core = core_tokens(name)
    if core and core in by_core:
        return by_core[core]
    nn = normname(name)
    if nn in by_name:
        return by_name[nn]
    return None


def register(o):
    if o.acro:
        by_acro.setdefault(o.acro, o)
    if o.core:
        by_core.setdefault(o.core, o)
    by_name.setdefault(normname(o.name), o)


def upsert(name, acro, *, subtype=None, email=None, phone=None, contact_name=None,
           website=None, notes=None):
    name = clean_name(name)
    if not name:
        return None
    o = find(name, acro)
    if o is None:
        o = Org()
        o.name = name
        o.core = core_tokens(name)
        orgs.append(o)
        register(o)
    # prefer the longer, fuller name
    if len(name) > len(o.name or ""):
        o.name = name
        o.core = core_tokens(name)
        register(o)
    if acro and not o.acro:
        o.acro = acro
        by_acro.setdefault(acro, o)
    if subtype and not o.subtype:
        o.subtype = subtype
    for attr, val in (("email", email), ("phone", phone),
                      ("contact_name", contact_name), ("website", website)):
        if val and not getattr(o, attr):
            setattr(o, attr, val)
    for n in (notes or []):
        n = cell(n)
        if n and n not in o.notes:
            o.notes.append(n)
    return o


def subtype_for(name, section):
    n = name.lower()
    if any(k in n for k in ("im3", "dentalaire", "midmark")):
        return "Equipment Vendor"
    if any(k in n for k in ("news letter", "newsletter", "dvm360", "clinician")):
        return "Industry Media"
    if any(k in n for k in ("conference", "forum", "fetch", "intro vet", "pacvet",
                            "pacific veterinary")):
        return "Conference / CE"
    if "board" in n:
        return "Veterinary Board"
    if section == "alumni":
        return "Alumni Group"
    if section == "facebook":
        return "Facebook Group"
    return "Veterinary Association"


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    # ---- Sheet 1: CE Research AssociationsGroupsE -------------------------
    ws = wb["CE Research AssociationsGroupsE"]
    section = "assoc"
    for row in ws.iter_rows(values_only=True):
        c = [cell(x) for x in row] + [""] * 5
        name = c[0]
        upper = name.upper()
        if not name:
            continue
        if "LOCAL CLINICS" in upper:
            break  # stop — clinics belong to the referral CRM, not vendor
        if upper.startswith("CE RESEARCH"):
            if "ALUMNI" in upper:
                section = "alumni"
            elif "FACEBOOK" in upper:
                section = "facebook"
            elif "ASSOCIATION" in upper or "GROUP" in upper:
                section = "assoc"
            continue
        if name.startswith("Name:"):
            continue
        if upper in ("SOUTH BAY", "RIVERSIDE", "OUTSIDE OF THE BOX"):
            continue
        email, email_note = first_email(c[1])
        phone, phone_note = first_phone(c[2])
        director = c[3] or None
        # col4 may be an acronym OR a website
        acro = clean_acro(c[4])
        website = first_url(c[4]) if not acro else None
        notes = []
        if email_note:
            notes.append(email_note)
        if phone_note:
            notes.append(phone_note)
        if c[4] and not acro and not website and c[4].lower() not in (
                "face book group", "facebook group"):
            notes.append(c[4])
        upsert(
            name, acro, subtype=subtype_for(name, section), email=email,
            phone=phone, contact_name=director, website=website, notes=notes,
        )

    # ---- Sheet 2: Local Veterinary Groups (acronyms + descriptions) -------
    ws = wb["Local Veterinary Groups"]
    skip_headers = {"local veterinary groups", "outside of the box",
                    "other communities"}
    for row in ws.iter_rows(values_only=True):
        c = [cell(x) for x in row] + [""] * 5
        name = c[0]
        if not name or name.lower() in skip_headers:
            continue
        if c[0].startswith("Name of Veterinary"):
            continue
        acro = clean_acro(c[1])
        info = c[4]
        website = first_url(info)
        notes = [info] if info else []
        # default subtype if this group is new (not matched to sheet1)
        sub = subtype_for(name, "assoc")
        upsert(name, acro, subtype=sub, website=website, notes=notes)

    wb.close()

    # ---- emit SQL ----------------------------------------------------------
    def s(v):
        if v is None or v == "":
            return "null"
        text = re.sub(r"\s+", " ", str(v)).strip()
        if not text:
            return "null"
        return "'" + text.replace("'", "''") + "'"

    out = ["set search_path = greendogops, public;"]
    n = 0
    for o in orgs:
        if not o.name:
            continue
        n += 1
        display = o.name
        if o.acro and o.acro.lower() not in normname(o.name) + " " \
                and f"({o.acro})" not in display:
            display = f"{o.name} ({o.acro})"
        notes = list(o.notes)
        notes_str = " | ".join(notes) if notes else None
        out.append(
            "insert into greendogops.crm_organization "
            "(org_type, name, subtype, contact_name, phone, email, website, "
            "notes, status, source) "
            "select 'med_ops', "
            f"{s(display)}, {s(o.subtype or 'Veterinary Association')}, "
            f"{s(o.contact_name)}, {s(o.phone)}, {s(o.email)}, {s(o.website)}, "
            f"{s(notes_str)}, 'lead', '{SOURCE_TAG}' "
            "where not exists (select 1 from greendogops.crm_organization c "
            "where regexp_replace(lower(c.name), '[^a-z0-9]', '', 'g') = "
            f"regexp_replace(lower({s(display)}), '[^a-z0-9]', '', 'g'));"
        )
    out.append(
        "select count(*) filter (where source = '" + SOURCE_TAG + "') as imported "
        "from greendogops.crm_organization;"
    )
    sys.stdout.write("\n".join(out) + "\n")
    sys.stderr.write(f"parsed {n} associations/groups from {SRC}\n")


if __name__ == "__main__":
    main()
