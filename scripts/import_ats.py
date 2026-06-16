#!/usr/bin/env python3
"""Parse the 2026 recruiting grid (multi-tab xlsx) into greendogops.person +
person_recruiting inserts. Skips divider/junk rows and dedupes across tabs."""
import re
import sys
from datetime import datetime, date

import openpyxl

SRC = ".data/2026 NEW GD RECRUITING GRID.xlsx"

# Tabs that hold real candidate pipelines -> (sheet title, header_row_index)
PIPELINES = {
    "All In House Positions": 1,
    "Remote CSR": 1,
    "DVM Vet America ": 0,
    "Volunteers": 0,
    "Externs": 0,
}
HIRED_TAB = "HIRED "

# Map raw header text (lowercased, stripped) -> canonical field
HEADER_MAP = {
    "applicant": "name",
    "position": "target_title",
    "status notes": "status_notes",
    "stage": "stage",
    "email": "email",
    "phone": "phone",
    "interview date:": "interview_date",
    "date of interview": "interview_date",
    "date of interview:": "interview_date",
    "score": "score",
    "notes": "notes",
    "resume": "resume",
    "csr exercise responses": "exercise",
    "found on:": "source",
    "keep for future": "keep_for_future",
    "follow up date": "follow_up_date",
    "follow up date:": "follow_up_date",
    "follow up date ": "follow_up_date",
    "start date:": "start_date",
    "end date:": "end_date",
}

# Junk applicant values to skip (dividers / placeholders)
JUNK_NAMES = {"team 3", "team 4", "team 5", "team 6", "others", "new", "inactive"}


def sql_str(v):
    if v is None:
        return "null"
    v = str(v).strip()
    if v == "" or v.lower() in ("none", "nan"):
        return "null"
    return "'" + v.replace("'", "''") + "'"


def sql_num(v):
    if v is None or str(v).strip() == "":
        return "null"
    try:
        return str(float(v))
    except (ValueError, TypeError):
        return "null"


def sql_bool(v):
    if v is None:
        return "null"
    t = str(v).strip().lower()
    if t in ("yes", "true", "1", "y"):
        return "true"
    if t in ("no", "false", "0", "n", ""):
        return "false" if t in ("no", "false", "0", "n") else "null"
    return "null"


def sql_date(v):
    if v is None:
        return "null"
    if isinstance(v, datetime):
        return "'" + v.strftime("%Y-%m-%d") + "'"
    if isinstance(v, date):
        return "'" + v.strftime("%Y-%m-%d") + "'"
    raw = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%b %d, %Y", "%B %d, %Y"):
        try:
            return "'" + datetime.strptime(raw, fmt).strftime("%Y-%m-%d") + "'"
        except ValueError:
            continue
    return "null"


def split_name(full):
    full = re.sub(r"\s+", " ", str(full).strip())
    full = full.strip("'\" ")
    if not full:
        return None, None
    parts = full.split(" ")
    if len(parts) == 1:
        return parts[0], None
    return parts[0], " ".join(parts[1:])


def is_junk(name, target_title):
    if not name:
        return True
    low = name.strip().lower()
    if low in JUNK_NAMES:
        return True
    # Divider rows: applicant text equals the position group and is ALL CAPS
    if target_title and name.strip() == str(target_title).strip() and name.isupper():
        return True
    if name.isupper() and len(name.split()) <= 3 and "@" not in name:
        # all-caps short label with no person feel -> likely a divider
        if not any(ch.islower() for ch in name):
            return True
    return False


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    candidates = []  # list of dicts
    seen = {}  # dedupe key -> index in candidates

    def add(rec):
        key = (rec.get("email") or "").strip().lower()
        if not key:
            fn = (rec.get("first_name") or "").lower()
            ln = (rec.get("last_name") or "").lower()
            key = f"{fn}|{ln}"
        if key in seen:
            # merge: keep existing, append pipeline if new
            existing = candidates[seen[key]]
            if rec["pipeline"] not in existing["pipeline"]:
                existing["pipeline"] += f", {rec['pipeline']}"
            return
        seen[key] = len(candidates)
        candidates.append(rec)

    # ---- standard pipeline tabs ----
    for title, hrow in PIPELINES.items():
        ws = wb[title]
        rows = list(ws.iter_rows(values_only=True))
        if hrow >= len(rows):
            continue
        headers = rows[hrow]
        colmap = {}
        for ci, h in enumerate(headers):
            if h is None:
                continue
            key = HEADER_MAP.get(str(h).strip().lower())
            if key and key not in colmap:
                colmap[key] = ci
        name_ci = colmap.get("name")
        if name_ci is None:
            continue
        for row in rows[hrow + 1:]:
            if name_ci >= len(row):
                continue
            raw_name = row[name_ci]
            tgt = row[colmap["target_title"]] if "target_title" in colmap and colmap["target_title"] < len(row) else None
            if is_junk(raw_name, tgt):
                continue
            fn, ln = split_name(raw_name)
            if not fn:
                continue

            def g(field):
                ci = colmap.get(field)
                if ci is None or ci >= len(row):
                    return None
                return row[ci]

            notes_extra = []
            for f in ("exercise", "start_date", "end_date"):
                v = g(f)
                if v:
                    notes_extra.append(f"{f}: {v}")
            base_notes = g("notes")
            full_notes = base_notes
            if notes_extra:
                full_notes = (str(base_notes) + " | " if base_notes else "") + " | ".join(notes_extra)

            add({
                "first_name": fn,
                "last_name": ln,
                "email": (str(g("email")).strip() if g("email") else None),
                "phone": g("phone"),
                "pipeline": title.strip(),
                "stage": g("stage"),
                "status_notes": g("status_notes"),
                "source": g("source"),
                "interview_date": g("interview_date"),
                "score": g("score"),
                "resume": g("resume"),
                "keep_for_future": g("keep_for_future"),
                "follow_up_date": g("follow_up_date"),
                "notes": full_notes,
                "target_title": tgt,
            })

    # ---- HIRED tab (different layout) ----
    ws = wb[HIRED_TAB]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    def hidx(name):
        for i, h in enumerate(headers):
            if h == name:
                return i
        return None
    ci_first = hidx("first name")
    ci_last = hidx("last name")
    ci_email = hidx("email")
    ci_phone = hidx("phone")
    ci_pos = hidx("position")
    ci_notes = hidx("notes / status")
    ci_found = hidx("found on")
    ci_loc = hidx("location")
    for row in rows[1:]:
        def gg(ci):
            return row[ci] if ci is not None and ci < len(row) else None
        fn = gg(ci_first)
        ln = gg(ci_last)
        if (not fn or not str(fn).strip()) and (not ln or not str(ln).strip()):
            continue
        nm = f"{fn or ''} {ln or ''}".strip()
        if is_junk(nm, None) or "@" in nm or "http" in nm.lower():
            continue
        if not fn and ln:
            fn, ln = split_name(ln)
        loc = gg(ci_loc)
        notes = gg(ci_notes)
        if loc:
            notes = (str(notes) + " | " if notes else "") + f"location: {loc}"
        add({
            "first_name": str(fn).strip() if fn else None,
            "last_name": str(ln).strip() if ln else None,
            "email": (str(gg(ci_email)).strip() if gg(ci_email) else None),
            "phone": gg(ci_phone),
            "pipeline": "Hired",
            "stage": "Hired",
            "status_notes": None,
            "source": gg(ci_found),
            "interview_date": None,
            "score": None,
            "resume": None,
            "keep_for_future": None,
            "follow_up_date": None,
            "notes": notes,
            "target_title": gg(ci_pos),
        })

    # ---- emit SQL ----
    out = [
        "set search_path = greendogops, public;",
        "-- ATS candidate import (idempotent-ish: clears prior import first)",
        "delete from greendogops.person_recruiting r using greendogops.person p "
        "where r.person_id = p.id and p.status = 'applicant' and p.created_by is null;",
        "delete from greendogops.person where status = 'applicant' and created_by is null;",
    ]
    for c in candidates:
        full = " ".join(x for x in [c["first_name"], c["last_name"]] if x)
        email_clean = c["email"]
        if email_clean and ("@" not in email_clean or " " in email_clean.strip()):
            # keep junky emails out of the email column, stash in notes
            c["notes"] = ((str(c["notes"]) + " | ") if c["notes"] else "") + f"email_raw: {email_clean}"
            email_clean = None
        out.append(
            "with p as (insert into greendogops.person "
            "(status, first_name, last_name, full_name, email, phone_mobile) values ("
            f"'applicant', {sql_str(c['first_name'])}, {sql_str(c['last_name'])}, "
            f"{sql_str(full)}, {sql_str(email_clean)}, {sql_str(c['phone'])}) returning id) "
            "insert into greendogops.person_recruiting "
            "(person_id, pipeline, stage, status_notes, source, interview_date, score, "
            "resume_url, keep_for_future, follow_up_date, notes, target_title) "
            f"select id, {sql_str(c['pipeline'])}, {sql_str(c['stage'])}, "
            f"{sql_str(c['status_notes'])}, {sql_str(c['source'])}, {sql_date(c['interview_date'])}, "
            f"{sql_num(c['score'])}, {sql_str(c['resume'])}, {sql_bool(c['keep_for_future'])}, "
            f"{sql_date(c['follow_up_date'])}, {sql_str(c['notes'])}, {sql_str(c['target_title'])} from p;"
        )
    out.append(
        "select count(*) as candidates from greendogops.person where status='applicant';"
    )
    sys.stdout.write("\n".join(out) + "\n")
    sys.stderr.write(f"Parsed {len(candidates)} unique candidates\n")


if __name__ == "__main__":
    main()
