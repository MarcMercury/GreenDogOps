#!/usr/bin/env python3
"""Parse the OLD MVS staff-recruitment grid into greendogops ATS records.

This is HISTORICAL recruiting data (≈2014-2021) from a prior grid. Unlike the
2026 importer this script is NON-DESTRUCTIVE: it never deletes existing rows.
Every candidate insert is guarded by a NOT EXISTS check against the live
greendogops.person table (matched by email, else first+last name), so:

  * existing candidates (e.g. the 320 from the 2026 grid) are left untouched,
  * duplicates across the MVS sheets are collapsed (statements run in order and
    later inserts see earlier ones), and
  * the script is safely re-runnable / idempotent.

Candidate-bearing sheets parsed:
  * 'NEW HIRES'      — structured applicant tracker (Inquiry -> Offer)
  * 'Staff Outreach' — sourced non-DVM leads (single NAME column)
  * 'Vet Outreach'   — sourced DVM leads (single NAME column)
The 'DUPS' sheet is used as an explicit skip-list of names the team already
flagged as duplicates. Process / job-board / business-lead sheets are ignored.
"""
import re
import sys
from datetime import datetime, date

import openpyxl

SRC = ".data/OLD - MVS STAFF RECRUITMENT.xlsx"
SOURCE_TAG = "OLD MVS grid"

CRED_SUFFIX_RE = re.compile(
    r"[,\s]+(dvm|rvt|cvt|vmd|dds|phd|bs|ba|ms|cva|cvpm|lvt)\.?$", re.IGNORECASE
)


# --------------------------------------------------------------------------- #
# SQL literal helpers
# --------------------------------------------------------------------------- #
def sql_str(v):
    if v is None:
        return "null"
    v = str(v).strip()
    if v == "" or v.lower() in ("none", "nan", "n/a"):
        return "null"
    return "'" + v.replace("'", "''") + "'"


def sql_num(v):
    if v is None or str(v).strip() == "":
        return "null"
    try:
        return str(float(v))
    except (ValueError, TypeError):
        return "null"


def sql_bool(v):
    if v is None:
        return "null"
    t = str(v).strip().lower()
    if t in ("yes", "true", "1", "y"):
        return "true"
    if t in ("no", "false", "0", "n"):
        return "false"
    return "null"


def sql_date(v):
    if v is None:
        return "null"
    if isinstance(v, (datetime, date)):
        return "'" + v.strftime("%Y-%m-%d") + "'"
    raw = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%b %d, %Y", "%B %d, %Y"):
        try:
            return "'" + datetime.strptime(raw, fmt).strftime("%Y-%m-%d") + "'"
        except ValueError:
            continue
    return "null"


# --------------------------------------------------------------------------- #
# parsing helpers
# --------------------------------------------------------------------------- #
def clean(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = re.sub(r"\s+", " ", str(v)).strip()
    s = re.sub(r"\s00:00:00$", "", s)
    return s or None


def clean_name(full):
    s = clean(full)
    if not s:
        return None
    s = s.strip("'\"")
    # strip a trailing credential suffix (", DVM", " RVT", ...)
    prev = None
    while prev != s:
        prev = s
        s = CRED_SUFFIX_RE.sub("", s).strip()
    return s or None


def split_name(full):
    s = clean_name(full)
    if not s:
        return None, None
    parts = s.split(" ")
    if len(parts) == 1:
        return parts[0], None
    return parts[0], " ".join(parts[1:])


def looks_like_person(name):
    """Reject dividers / junk / non-person cells."""
    if not name:
        return False
    low = name.lower()
    if "@" in low or "http" in low or "www." in low:
        return False
    if not any(ch.isalpha() for ch in name):
        return False
    # all-caps short label with no lowercase -> section divider
    if name.isupper() and not any(ch.islower() for ch in name):
        return False
    junk = {
        "hired staff", "hired", "teeth cleaning companies", "name", "office",
        "new", "inactive", "others", "pass", "status", "new lead", "tbd",
    }
    if low in junk:
        return False
    return True


def derive_stage_from_status(status_text, default="Lead"):
    s = (status_text or "").lower()
    if not s:
        return default
    if "no longer with us" in s or "separat" in s or "quit" in s or "fired" in s:
        return "No Longer With Us"
    if "still with us" in s or ("hired" in s and "not hired" not in s):
        return "Hired"
    if "extern" in s:
        return "Externship"
    if "no response" in s or "did not respond" in s:
        return "No Response"
    if "pass" in s or "decline" in s or " no " == f" {s.strip()} " or s.strip() == "no":
        return "Pass"
    if "hold" in s or "future" in s:
        return "Hold for Future"
    if "interview" in s and "pass" not in s:
        return "Interviewed"
    if "in person" in s or "shadow" in s or "ipi" in s or "f2f" in s:
        return "In Person / Shadow"
    if "phoner" in s or "phone" in s:
        return "Phone Screen"
    if "reached out" in s or "interested" in s or "new lead" in s:
        return "New Lead"
    return default


# --------------------------------------------------------------------------- #
# candidate collection
# --------------------------------------------------------------------------- #
candidates = []  # ordered list of dicts


def add(rec):
    fn = clean(rec.get("first_name"))
    ln = clean(rec.get("last_name"))
    if not fn:
        return
    full = clean(rec.get("full_name")) or " ".join(x for x in [fn, ln] if x)
    if not looks_like_person(full):
        return
    email = clean(rec.get("email"))
    if email and ("@" not in email or " " in email):
        # junky email -> stash in notes, don't pollute the email column
        rec["notes"] = ((rec.get("notes") + " | ") if rec.get("notes") else "") + f"email_raw: {email}"
        email = None
    rec["first_name"] = fn
    rec["last_name"] = ln
    rec["full_name"] = full
    rec["email"] = email
    candidates.append(rec)


def build_notes(*pairs):
    """pairs = (label, value) tuples; returns a ' | '-joined notes string."""
    out = []
    for label, value in pairs:
        v = clean(value)
        if v:
            out.append(f"{label}: {v}" if label else v)
    return " | ".join(out) if out else None


def parse_new_hires(wb, dup_skip):
    ws = wb["NEW HIRES"]
    rows = list(ws.iter_rows(values_only=True))

    def g(row, i):
        return row[i] if i < len(row) else None

    for row in rows[1:]:
        fn = clean(g(row, 2))
        ln = clean(g(row, 3))
        if not fn and not ln:
            continue
        if not fn:
            fn, ln = split_name(ln)
        if (fn or "", ln or "") and (fn.lower() if fn else "", (ln or "").lower()) in dup_skip:
            continue
        offer = clean(g(row, 24))
        start = g(row, 26)
        iview_result = clean(g(row, 18))
        iview_date = g(row, 16)
        if offer and offer.lower().startswith("y") and start:
            stage = "Hired"
        elif offer and offer.lower().startswith("y"):
            stage = "Offer"
        elif iview_result and iview_result.lower().startswith("y"):
            stage = "Interviewed"
        elif iview_date:
            stage = "Interviewed"
        else:
            stage = "Applicant"
        notes = build_notes(
            (None, g(row, 21)),
            ("Address", g(row, 7)),
            ("Pay offer", g(row, 25)),
            ("Start date", clean(str(start)) if start else None),
            ("Contact method", g(row, 11)),
        )
        add({
            "first_name": fn,
            "last_name": ln,
            "email": g(row, 4),
            "phone": clean(g(row, 5)),
            "postal_code": clean(str(g(row, 9)).replace(".0", "")) if g(row, 9) else None,
            "pipeline": "MVS New Hires",
            "stage": stage,
            "status_notes": None,
            "source": clean(g(row, 1)) or SOURCE_TAG,
            "interview_date": iview_date,
            "score": None,
            "follow_up_date": g(row, 0),  # inquiry date as a historical anchor
            "notes": notes,
            "target_title": clean(g(row, 6)),
        })


def parse_outreach(wb, sheet, dup_skip, *, name_col, status_col, phone_col,
                   email_col, pos_col, pipeline, target_default=None,
                   extra_note_cols=()):
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))

    def g(row, i):
        return row[i] if i is not None and i < len(row) else None

    for row in rows[2:]:  # row 0 = column-group labels, row 1 = headers
        raw_name = g(row, name_col)
        fn, ln = split_name(raw_name)
        if not fn:
            continue
        if (fn.lower(), (ln or "").lower()) in dup_skip:
            continue
        status = clean(g(row, status_col))
        note_pairs = [(label, g(row, ci)) for label, ci in extra_note_cols]
        add({
            "first_name": fn,
            "last_name": ln,
            "email": g(row, email_col),
            "phone": clean(g(row, phone_col)),
            "postal_code": None,
            "pipeline": pipeline,
            "stage": derive_stage_from_status(status),
            "status_notes": status,
            "source": SOURCE_TAG,
            "interview_date": None,
            "score": None,
            "follow_up_date": None,
            "notes": build_notes(*note_pairs),
            "target_title": clean(g(row, pos_col)) if pos_col is not None else target_default,
        })


def load_dup_skip(wb):
    """Names the team already flagged as duplicates -> set of (first, last)."""
    ws = wb["DUPS"]
    skip = set()
    for row in ws.iter_rows(values_only=True):
        c0 = clean(row[0]) if len(row) > 0 else None
        c1 = clean(row[1]) if len(row) > 1 else None
        if c0 and c1:
            skip.add((c0.lower(), c1.lower()))
        elif c0:
            fn, ln = split_name(c0)
            if fn:
                skip.add((fn.lower(), (ln or "").lower()))
    return skip


# --------------------------------------------------------------------------- #
# main
# --------------------------------------------------------------------------- #
def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    dup_skip = load_dup_skip(wb)

    parse_new_hires(wb, dup_skip)
    parse_outreach(
        wb, "Staff Outreach", dup_skip,
        name_col=5, status_col=1, phone_col=6, email_col=7, pos_col=3,
        pipeline="MVS Staff Outreach",
        extra_note_cols=[
            ("Location", 8), ("Phone notes", 2), ("Site found", 4),
            ("Found", 10), ("Experience", 14), ("FB", 9), ("LinkedIn", 11),
            ("Instagram", 12), ("Website", 13), ("Intro email", 15),
        ],
    )
    parse_outreach(
        wb, "Vet Outreach", dup_skip,
        name_col=4, status_col=1, phone_col=5, email_col=6, pos_col=None,
        pipeline="MVS Vet Outreach", target_default="DVM",
        extra_note_cols=[
            ("Availability", 2), ("Rate", 3), ("Location", 7), ("Source", 8),
            ("LinkedIn", 9), ("Practice site", 10), ("Personal site", 11),
            ("Notes", 12), ("Skype", 13),
        ],
    )
    wb.close()

    out = [
        "set search_path = greendogops, public;",
        "-- OLD MVS recruiting import — NON-DESTRUCTIVE, deduped via NOT EXISTS.",
    ]
    for c in candidates:
        email_lit = sql_str(c["email"])
        fn_lit = sql_str(c["first_name"])
        ln_lit = sql_str(c["last_name"])
        full_lit = sql_str(c["full_name"])
        out.append(
            "with p as (\n"
            "  insert into greendogops.person\n"
            "    (status, first_name, last_name, full_name, email, phone_mobile, postal_code)\n"
            f"  select 'applicant', {fn_lit}, {ln_lit}, {full_lit}, {email_lit}, "
            f"{sql_str(c['phone'])}, {sql_str(c['postal_code'])}\n"
            "  where not exists (\n"
            "    select 1 from greendogops.person ex where\n"
            f"      ({email_lit} is not null and ex.email is not null "
            f"and lower(ex.email) = lower({email_lit}))\n"
            f"      or ({fn_lit} is not null and lower(coalesce(ex.first_name,'')) = lower(coalesce({fn_lit},''))\n"
            f"          and lower(coalesce(ex.last_name,'')) = lower(coalesce({ln_lit},'')))\n"
            "  )\n"
            "  returning id\n"
            ")\n"
            "insert into greendogops.person_recruiting\n"
            "  (person_id, pipeline, stage, status_notes, source, interview_date,\n"
            "   score, keep_for_future, follow_up_date, notes, target_title)\n"
            f"select id, {sql_str(c['pipeline'])}, {sql_str(c['stage'])}, "
            f"{sql_str(c['status_notes'])}, {sql_str(c['source'])}, {sql_date(c['interview_date'])},\n"
            f"  {sql_num(c['score'])}, null, {sql_date(c['follow_up_date'])}, "
            f"{sql_str(c['notes'])}, {sql_str(c['target_title'])} from p;"
        )
    out.append("select count(*) as total_applicants from greendogops.person where status='applicant';")
    sys.stdout.write("\n".join(out) + "\n")
    sys.stderr.write(
        f"Parsed {len(candidates)} candidate rows from MVS grid "
        f"(skip-list had {len(dup_skip)} names)\n"
    )


if __name__ == "__main__":
    main()
